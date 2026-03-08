from __future__ import annotations

import csv
import hashlib
import json
import logging
import re
import sqlite3
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Iterator

try:
    import fiona
except ModuleNotFoundError:  # pragma: no cover - exercised in minimal test envs
    fiona = None

try:
    import openpyxl
except ModuleNotFoundError:  # pragma: no cover - exercised in minimal test envs
    openpyxl = None


LOG = logging.getLogger("service_ref")

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "service_ref" / "output"
DB_PATH = OUT_DIR / "service_referential.sqlite"

LEA_PATH = ROOT / "6-3_20260203_Suivi_Contrats_LEA.xlsx"
GDB_ZIP_PATH = ROOT / "GDB_TeloiseV3 (1).zip"
SWAG_PATH = ROOT / "unzipped_equip" / "Export inventaire SWAG.xlsx"
CPE_PATH = ROOT / "unzipped_equip" / "Inventaire CPE Teloise Janv26.xlsx"
CONFIG_DIR = ROOT / "unzipped_equip"

GDB_URI = (
    f"zip://{GDB_ZIP_PATH.as_posix()}!TELOISE_TELOISE_20250610-130725.gdb"
)

ROLE_BY_OFFER = {
    "fibre-longue distance-iru": ("IRU FON", "principal"),
    "fibre-longue distance-maintenance": ("IRU FON", "maintenance"),
    "fibre-metro-iru": ("IRU FON", "principal"),
    "fibre-metro-maintenance": ("IRU FON", "maintenance"),
    "fibre-longue distance-location": ("Location FON", "principal"),
    "fibre-metro-location": ("Location FON", "principal"),
    "fibre-metro-location escomptee": ("Location FON", "principal"),
    "lien ethernet - 1-100 mbits": ("Lan To Lan", "principal"),
    "lien sdh - e1 (2 mbits)": ("Lan To Lan", "principal"),
    "netcenter lan - divers": ("Lan To Lan", "annexe"),
    "netcenter lan - extra works": ("Lan To Lan", "frais_acces"),
    "emplacement baie - 48v - 1, 3 ou 5 ans": ("Hebergement", "principal"),
    "netcenter baie - divers": ("Hebergement", "annexe"),
    "fibre - extra works": ("Hebergement", "frais_acces"),
}

ROUTE_PATTERN = re.compile(
    r"\b(TOIP\s*\d+|00FT\s*\d+|FREE\s*\d+|SFRH\d+|OPSC\s*\d+|9CIP\s*\d+|COMP\s*\d+|BOUY\s*\d+)\b",
    re.IGNORECASE,
)
SERVICE_PATTERN = re.compile(
    r"\b((?:OPE|PE)\s*\d+)\s*[/_-]?\s*(L2L\s*\d+)\b",
    re.IGNORECASE,
)
VLAN_LINE_PATTERN = re.compile(r"\bvlan\s+(\d+)\b", re.IGNORECASE)
HOSTNAME_PATTERN = re.compile(r"\bsysname\s+([^\s]+)|\bhostname\s+([^\s]+)", re.IGNORECASE)
HEADER_PATTERN = re.compile(r'header shell information\s+"([^"]+)"', re.IGNORECASE)
CABLE_FIBERS_PATTERN = re.compile(r"\b(\d{1,3})\s*FO\b", re.IGNORECASE)
PLACE_TOKEN_PATTERN = re.compile(r"[A-Z0-9]{3,}")
HOUSING_SITE_RELATION_LAYERS = {
    "Chamber__Hubsite": ("Chamber", "CHAMBER_MIGRATION_OID", "HUBSITE_MIGRATION_OID"),
    "SupportStructure__Hubsite": ("SupportStructure", "SUPPORTSTRUCTURE_MIGRATION_OID", "HUBSITE_MIGRATION_OID"),
}


@dataclass
class SiteMatch:
    site_id: str | None
    score: int
    rule: str | None
    site_name: str | None


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")


def _require_dependency(module: object, package_name: str) -> object:
    if module is None:
        raise RuntimeError(
            f"Missing optional dependency '{package_name}'. Install project data dependencies "
            f"to run source-loading steps."
        )
    return module


def norm_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm_slug(value: object) -> str:
    return norm_text(value).replace(" ", "_")


def infer_route_ref(*values: object) -> str | None:
    refs = extract_route_refs(*values)
    return refs[0] if refs else None


def extract_place_tokens(*values: object) -> list[str]:
    tokens: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = norm_text(value)
        if not normalized:
            continue
        for token in PLACE_TOKEN_PATTERN.findall(normalized):
            if (
                token.isdigit()
                or len(token) < 4
                or token in BUSINESS_STOPWORDS
                or token.endswith("FO")
            ):
                continue
            if token not in seen:
                tokens.append(token)
                seen.add(token)
    return tokens


def extract_fiber_count(*values: object) -> int | None:
    for value in values:
        text = "" if value is None else str(value)
        match = CABLE_FIBERS_PATTERN.search(text)
        if match:
            try:
                return int(match.group(1))
            except ValueError:
                continue
    return None


ROLE_BY_OFFER_NORMALIZED = {norm_text(key): value for key, value in ROLE_BY_OFFER.items()}


def extract_route_refs(*values: object) -> list[str]:
    refs: list[str] = []
    for value in values:
        text = "" if value is None else str(value)
        refs.extend(" ".join(match.split()) for match in ROUTE_PATTERN.findall(text))
    deduped: list[str] = []
    seen = set()
    for ref in refs:
        ref = ref.upper().strip()
        if ref not in seen:
            deduped.append(ref)
            seen.add(ref)
    return deduped


def extract_service_refs(*values: object) -> list[str]:
    refs: list[str] = []
    for value in values:
        text = "" if value is None else str(value)
        for left, right in SERVICE_PATTERN.findall(text):
            refs.append(f"{norm_text(left).replace(' ', '')}/{norm_text(right).replace(' ', '')}")
    deduped: list[str] = []
    seen = set()
    for ref in refs:
        if ref not in seen:
            deduped.append(ref)
            seen.add(ref)
    return deduped


BUSINESS_STOPWORDS = {
    "CLIENT",
    "CLIENTS",
    "LAN2LAN",
    "L2L",
    "TRUNK",
    "VERS",
    "CPE",
    "DSP",
    "VLAN",
    "POP",
    "SITE",
    "PORT",
    "ACCES",
    "ACCESS",
    "TRANSPORT",
    "COLLECTE",
    "SERVICE",
    "SERVICES",
    "SHUT",
    "TEMPORAIRE",
    "NE",
    "PAS",
    "CPEDSP",
    "CLIENTLAN2LAN",
    "CLIENTL2L",
    "TELOISE",
    "BAIE",
    "HOTEL",
    "POLICE",
    "RUE",
}

INTERFACE_START_RE = re.compile(r"^interface\s+(.+)$", re.IGNORECASE)
RAD_ETHERNET_START_RE = re.compile(r"^ethernet\s+(\d+)$", re.IGNORECASE)
VLAN_START_RE = re.compile(r"^vlan\s+(\d+)$", re.IGNORECASE)
DESCRIPTION_RE = re.compile(r'^(?:description\s+|name\s+")(.+?)(?:"$)?$', re.IGNORECASE)
ANNOUNCEMENT_RE = re.compile(r"announcement\s+'([^']+)'", re.IGNORECASE)
LOCATION_RE = re.compile(r'location\s+"([^"]+)"', re.IGNORECASE)
RAD_SYSTEM_NAME_RE = re.compile(r'^name\s+"([^"]+)"$', re.IGNORECASE)
HEADER_CLIENT_SITE_RE = re.compile(r"CLIENT\s*:\s*(.*?)\s+SITE\s*:\s*(.*?)\s+DSP\s*:", re.IGNORECASE)


def parse_vlan_list(text: str) -> list[int]:
    values: list[int] = []
    cleaned = re.sub(r".*?vlan(?:\s+add)?\s+", "", text, flags=re.IGNORECASE)
    cleaned = cleaned.replace("to", "-")
    for part in re.split(r"[,\s]+", cleaned):
        token = part.strip()
        if not token:
            continue
        if "-" in token:
            left, right = token.split("-", 1)
            if left.isdigit() and right.isdigit():
                start = int(left)
                end = int(right)
                if start <= end and end - start <= 512:
                    values.extend(range(start, end + 1))
        elif token.isdigit():
            values.append(int(token))
    deduped = []
    seen = set()
    for value in values:
        if value not in seen:
            deduped.append(value)
            seen.add(value)
    return deduped


def extract_vlans_from_line(line: str) -> list[int]:
    lowered = line.lower()
    if any(
        marker in lowered
        for marker in [
            "switchport trunk allowed vlan",
            "port trunk allow-pass vlan",
            "port hybrid tagged vlan",
            "port hybrid untagged vlan",
            "port hybrid pvid vlan",
            "port default vlan",
            "port access vlan",
        ]
    ):
        return parse_vlan_list(line)
    return []


def clean_business_label(value: object) -> str:
    label = norm_text(value)
    if not label:
        return ""
    label = re.sub(r"\bTOIP\s*\d+\b", " ", label)
    label = re.sub(r"\b(?:OPE|PE)\s*\d+\b", " ", label)
    label = re.sub(r"\bL2L\s*\d+\b", " ", label)
    label = re.sub(r"\b\d{3,5}\b", " ", label)
    tokens = [token for token in label.split() if token not in BUSINESS_STOPWORDS]
    return " ".join(tokens).strip()


def business_tokens(value: object) -> set[str]:
    return {token for token in clean_business_label(value).split() if len(token) >= 4}


def detect_vendor(text: str) -> str:
    header = text[:500].upper()
    if "RANCID-CONTENT-TYPE: HUAWEI" in header or "HWTACACS" in header:
        return "huawei"
    if header.startswith("VERSION ") or "HOSTNAME " in header:
        return "cisco"
    if "ANNOUNCEMENT '-----CPE L2 RAD" in text.upper() or "ETX" in header:
        return "rad"
    return "unknown"


def detect_source_family(path: Path, device_name: str) -> str:
    name = path.name.upper()
    device = device_name.upper()
    if "_CO_" in name or "-CO-" in device:
        return "co"
    if "_SWS_" in name or "-SWS-" in device:
        return "sws"
    if "_SW_" in name or "-SW-" in device:
        return "sw"
    if "_SEC_" in name or "-SEC-" in device:
        return "sec"
    if "EXEMPLE" in name:
        return "cpe_example"
    return "config"


def extract_client_site_from_header(header_info: str) -> tuple[str, str]:
    match = HEADER_CLIENT_SITE_RE.search(header_info or "")
    if not match:
        return "", ""
    return match.group(1).strip(), match.group(2).strip()


def load_network_text_artifacts(con: sqlite3.Connection) -> None:
    LOG.info("Loading network text artifacts")
    device_records = []
    interface_records = []
    vlan_records = []
    device_counter = 0
    interface_counter = 0
    vlan_counter = 0

    for path in sorted(CONFIG_DIR.rglob("*.txt")):
        payload = path.read_text(encoding="latin1", errors="ignore")
        hostname = ""
        match = HOSTNAME_PATTERN.search(payload)
        if match:
            hostname = next(group for group in match.groups() if group)
        if not hostname:
            system_match = RAD_SYSTEM_NAME_RE.search(payload)
            if system_match:
                hostname = system_match.group(1)
        if not hostname:
            hostname = path.stem

        vendor = detect_vendor(payload)
        source_family = detect_source_family(path, hostname)
        announcement = ""
        announcement_match = ANNOUNCEMENT_RE.search(payload)
        if announcement_match:
            announcement = announcement_match.group(1).strip()
        header_match = HEADER_PATTERN.search(payload)
        header_info = header_match.group(1).strip() if header_match else ""
        header_client, header_site = extract_client_site_from_header(header_info)
        location_match = LOCATION_RE.search(payload)
        location_text = location_match.group(1).strip() if location_match else ""
        site_hint = header_site or location_text
        client_hint = header_client
        if announcement and not site_hint:
            site_from_announcement = re.search(r"SITE\s+(.+?)--", announcement, re.IGNORECASE)
            site_hint = site_from_announcement.group(1).strip() if site_from_announcement else ""
        if announcement and not client_hint:
            client_from_announcement = re.search(r"L2 RAD\s+\S+\s+(.+?)--SITE", announcement, re.IGNORECASE)
            client_hint = client_from_announcement.group(1).strip() if client_from_announcement else ""

        service_refs = extract_service_refs(payload, announcement, header_info)
        route_refs = extract_route_refs(payload, announcement, header_info)
        all_vlans = sorted({int(vlan) for vlan in VLAN_LINE_PATTERN.findall(payload)})
        device_counter += 1
        device_id = f"NDV-{device_counter:06d}-{safe_hash([str(path.relative_to(ROOT)), hostname])}"
        device_records.append(
            (
                device_id,
                hostname,
                str(path.relative_to(ROOT)),
                source_family,
                vendor,
                location_text,
                announcement,
                header_info,
                client_hint,
                site_hint,
                json.dumps(service_refs, ensure_ascii=True),
                json.dumps(route_refs, ensure_ascii=True),
                json.dumps(all_vlans, ensure_ascii=True),
                norm_text(" ".join([hostname, announcement, header_info, location_text])),
            )
        )

        current_interface_name = None
        current_interface_desc: list[str] = []
        current_interface_vlans: list[int] = []
        current_vlan_id: int | None = None
        current_vlan_label = ""

        def flush_interface() -> None:
            nonlocal current_interface_name, current_interface_desc, current_interface_vlans, interface_counter
            if not current_interface_name:
                return
            interface_counter += 1
            description = " | ".join(part for part in current_interface_desc if part).strip()
            interface_records.append(
                (
                    f"NIF-{interface_counter:07d}-{safe_hash([device_id, current_interface_name, description])}",
                    device_id,
                    hostname,
                    str(path.relative_to(ROOT)),
                    source_family,
                    current_interface_name,
                    description,
                    json.dumps(extract_service_refs(current_interface_name, description), ensure_ascii=True),
                    json.dumps(extract_route_refs(current_interface_name, description), ensure_ascii=True),
                    json.dumps(sorted(set(current_interface_vlans)), ensure_ascii=True),
                    clean_business_label(description or current_interface_name),
                )
            )
            current_interface_name = None
            current_interface_desc = []
            current_interface_vlans = []

        def flush_vlan() -> None:
            nonlocal current_vlan_id, current_vlan_label, vlan_counter
            if current_vlan_id is None:
                return
            vlan_counter += 1
            vlan_records.append(
                (
                    f"NVL-{vlan_counter:07d}-{safe_hash([device_id, current_vlan_id, current_vlan_label])}",
                    device_id,
                    hostname,
                    str(path.relative_to(ROOT)),
                    source_family,
                    current_vlan_id,
                    current_vlan_label,
                    json.dumps(extract_service_refs(current_vlan_label), ensure_ascii=True),
                    json.dumps(extract_route_refs(current_vlan_label), ensure_ascii=True),
                    clean_business_label(current_vlan_label),
                )
            )
            current_vlan_id = None
            current_vlan_label = ""

        for raw_line in payload.splitlines():
            line = raw_line.rstrip()
            stripped = line.strip()
            interface_match = INTERFACE_START_RE.match(stripped)
            rad_interface_match = RAD_ETHERNET_START_RE.match(stripped)
            vlan_match = VLAN_START_RE.match(stripped)
            if interface_match:
                flush_interface()
                flush_vlan()
                current_interface_name = interface_match.group(1).strip()
                continue
            if rad_interface_match:
                flush_interface()
                flush_vlan()
                current_interface_name = f"ethernet {rad_interface_match.group(1)}"
                continue
            if vlan_match and not current_interface_name:
                flush_interface()
                flush_vlan()
                current_vlan_id = int(vlan_match.group(1))
                continue
            if stripped in {"#", "!", "exit"}:
                flush_interface()
                flush_vlan()
                continue

            if current_interface_name:
                desc_match = DESCRIPTION_RE.match(stripped)
                if stripped.lower().startswith("description "):
                    current_interface_desc.append(stripped.split(" ", 1)[1].strip())
                elif stripped.lower().startswith("name "):
                    current_interface_desc.append(stripped.split(" ", 1)[1].strip().strip('"'))
                elif desc_match and desc_match.group(1):
                    current_interface_desc.append(desc_match.group(1).strip())
                current_interface_vlans.extend(extract_vlans_from_line(stripped))
                continue

            if current_vlan_id is not None:
                if stripped.lower().startswith("description "):
                    current_vlan_label = stripped.split(" ", 1)[1].strip()
                elif stripped.lower().startswith("name "):
                    current_vlan_label = stripped.split(" ", 1)[1].strip().strip('"')

        flush_interface()
        flush_vlan()

    con.executemany("insert into ref_network_devices values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", device_records)
    con.executemany("insert into ref_network_interfaces values (?,?,?,?,?,?,?,?,?,?,?)", interface_records)
    con.executemany("insert into ref_network_vlans values (?,?,?,?,?,?,?,?,?,?)", vlan_records)
    con.commit()
    LOG.info(
        "Loaded %s network devices, %s interfaces, %s vlan entries",
        len(device_records),
        len(interface_records),
        len(vlan_records),
    )


def build_party_master(con: sqlite3.Connection) -> None:
    LOG.info("Building party master")
    party_sources: list[tuple[str, str, str, str, int]] = []

    for row in con.execute("select distinct client_contractant from lea_active_lines where trim(client_contractant) <> ''"):
        party_sources.append((row[0], row[0], "contractant", "lea_active_lines", 100))
    for row in con.execute("select distinct client_final from lea_active_lines where trim(client_final) <> ''"):
        party_sources.append((row[0], row[0], "client_final", "lea_active_lines", 95))
    for row in con.execute("select distinct label from ref_network_vlans where trim(label) <> ''"):
        label = clean_business_label(row[0])
        if label:
            party_sources.append((label, row[0], "network_vlan", "ref_network_vlans", 70))
    for row in con.execute("select distinct description from ref_network_interfaces where trim(description) <> ''"):
        label = clean_business_label(row[0])
        if label:
            party_sources.append((label, row[0], "network_interface", "ref_network_interfaces", 65))
    for row in con.execute("select distinct client_hint from ref_network_devices where trim(client_hint) <> ''"):
        label = clean_business_label(row[0]) or row[0]
        party_sources.append((label, row[0], "device_client_hint", "ref_network_devices", 75))

    priority_order = {"contractant": 1, "client_final": 2, "device_client_hint": 3, "network_vlan": 4, "network_interface": 5}
    by_norm: dict[str, list[tuple[str, str, str, str, int]]] = defaultdict(list)
    for canonical_candidate, alias_value, party_type, source_table, source_priority in party_sources:
        normalized = norm_text(canonical_candidate)
        if normalized:
            by_norm[normalized].append((canonical_candidate, alias_value, party_type, source_table, source_priority))

    party_records = []
    alias_records = []
    alias_counter = 0
    for normalized, items in by_norm.items():
        items.sort(key=lambda item: (priority_order.get(item[2], 99), -item[4], len(norm_text(item[0]))))
        canonical_name, _, party_type, _, source_priority = items[0]
        party_id = f"PTY-{safe_hash([normalized])}"
        party_records.append((party_id, canonical_name, normalized, party_type, source_priority))
        for canonical_candidate, alias_value, alias_type, source_table, _ in items:
            alias_counter += 1
            alias_id = f"PAL-{alias_counter:07d}-{safe_hash([party_id, alias_value, alias_type, source_table])}"
            alias_records.append((alias_id, party_id, alias_value, norm_text(alias_value), source_table, alias_type))

    con.executemany("insert into party_master values (?,?,?,?,?)", party_records)
    con.executemany("insert into party_alias values (?,?,?,?,?,?)", alias_records)
    con.commit()
    LOG.info("Built %s parties and %s aliases", len(party_records), len(alias_records))


def safe_hash(parts: Iterable[object]) -> str:
    payload = "||".join(norm_text(part) for part in parts)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()[:16]


def classify_offer(offer: object) -> tuple[str, str]:
    normalized = norm_text(offer)
    if normalized in ROLE_BY_OFFER_NORMALIZED:
        return ROLE_BY_OFFER_NORMALIZED[normalized]
    if "MAINTENANCE" in normalized:
        return ("IRU FON", "maintenance")
    if "EXTRA WORKS" in normalized:
        return ("Lan To Lan", "frais_acces")
    return ("A qualifier", "principal")


def create_schema(con: sqlite3.Connection) -> None:
    con.executescript(
        """
        drop table if exists lea_active_lines;
        drop table if exists ref_sites;
        drop table if exists ref_optical_logical_route;
        drop table if exists ref_optical_lease;
        drop table if exists ref_optical_lease_endpoint;
        drop table if exists ref_optical_cable;
        drop table if exists ref_optical_housing;
        drop table if exists ref_optical_connection;
        drop table if exists ref_optical_site_link;
        drop table if exists ref_optical_cable_site_hint;
        drop table if exists ref_routes;
        drop table if exists ref_route_parcours;
        drop table if exists ref_lease_template;
        drop table if exists ref_fiber_lease;
        drop table if exists ref_isp_lease;
        drop table if exists ref_swag_interfaces;
        drop table if exists ref_cpe_inventory;
        drop table if exists ref_cpe_configs;
        drop table if exists ref_network_devices;
        drop table if exists ref_network_interfaces;
        drop table if exists ref_network_vlans;
        drop table if exists party_master;
        drop table if exists party_alias;
        drop table if exists service_master_active;
        drop table if exists service_bss_line;
        drop table if exists service_party;
        drop table if exists service_endpoint;
        drop table if exists service_support_optique;
        drop table if exists service_support_reseau;
        drop table if exists service_match_evidence;
        drop table if exists service_review_queue;
        drop table if exists gold_service_active;
        drop table if exists service_facturable_final;
        drop table if exists override_party_alias;
        drop table if exists override_site_alias;
        drop table if exists override_service_match;

        create table lea_active_lines (
            lea_line_id text primary key,
            client_contractant text,
            offer_label text,
            nature_service text,
            role_ligne text,
            command_internal text,
            command_external text,
            contract_file text,
            endpoint_a_raw text,
            endpoint_z_raw text,
            client_final text,
            lineaire text,
            date_signature text,
            date_creation text,
            duree_service text,
            status_code text,
            date_livraison text,
            date_resiliation text,
            fms text,
            rm_initiale text,
            rm_derniere text,
            iru text,
            montant text,
            route_refs_json text,
            service_refs_json text,
            is_old integer,
            grouping_key text,
            source_row integer
        );

        create table ref_sites (
            site_id text primary key,
            migration_id text,
            reference text,
            userreference text,
            address1 text,
            function_code integer,
            reseau_code integer,
            manager_code integer,
            owner_code integer,
            precision_code integer,
            project_code text,
            normalized_reference text,
            normalized_userreference text,
            normalized_address text
        );

        create table ref_optical_logical_route (
            logical_route_id text primary key,
            route_ref text,
            source_layer text,
            source_oid text,
            ref_exploit text,
            reference text,
            path_id text,
            pair_oid text,
            feature_type text,
            lessee text,
            client text,
            network text,
            status text,
            comments text,
            normalized_label text
        );

        create table ref_optical_lease (
            optical_lease_id text primary key,
            lease_kind text,
            source_layer text,
            source_oid text,
            ref_exploit text,
            reference text,
            path_id text,
            pair_oid text,
            feature text,
            lessee text,
            client text,
            network text,
            status text,
            comments text,
            normalized_label text
        );

        create table ref_optical_lease_endpoint (
            optical_lease_id text,
            endpoint_label text,
            housing_type text,
            housing_migration_oid text,
            object_type text,
            object_migration_oid text,
            connector_start integer,
            connector_end integer,
            reference_label text,
            site_id text,
            site_name text,
            site_resolution_rule text,
            site_score integer
        );

        create table ref_optical_cable (
            cable_id text primary key,
            migration_oid text,
            migration_id text,
            reference text,
            userreference text,
            comments text,
            labeltext text,
            location text,
            number_of_fibers integer,
            cable_type text,
            network text,
            project_code text,
            status text,
            normalized_reference text,
            site_tokens_json text
        );

        create table ref_optical_housing (
            housing_id text primary key,
            migration_oid text,
            housing_type text,
            reference text,
            userreference text,
            description text,
            comments text,
            location text,
            parent_type text,
            parent_migration_oid text,
            site_id text,
            site_name text,
            site_resolution_rule text,
            site_score integer
        );

        create table ref_optical_connection (
            connection_id text primary key,
            housing_type text,
            housing_migration_oid text,
            obj1_type text,
            obj1_migration_oid text,
            obj1_connector1 integer,
            obj1_connector2 integer,
            obj2_type text,
            obj2_migration_oid text,
            obj2_connector1 integer,
            obj2_connector2 integer,
            branch_type text,
            tray_migration_id text,
            rattachement_id integer
        );

        create table ref_optical_site_link (
            object_type text,
            object_migration_oid text,
            site_id text,
            site_name text,
            rule_name text,
            score integer,
            source_layer text
        );

        create table ref_optical_cable_site_hint (
            cable_id text,
            site_token text,
            score integer,
            hint_rule text
        );

        create table ref_routes (
            route_id text primary key,
            route_ref text,
            network text,
            client text,
            lessee text,
            status text
        );

        create table ref_route_parcours (
            route_ref text,
            route_id text,
            step_no integer,
            step_type text,
            site text,
            site_detail text,
            address text,
            bpe text,
            cable_in text,
            cable_out text,
            fibre_in text,
            fibre_out text,
            commentaire text
        );

        create table ref_lease_template (
            lease_id text primary key,
            ref_exploit text,
            reseau text,
            lessee text,
            clients text,
            housing_type_l1 text,
            housing_migoid_l1 text,
            type_l1 text,
            migoid_l1 text,
            l1_conn1 integer,
            l1_conn2 integer,
            reference_l1 text,
            housing_type_l2 text,
            housing_migoid_l2 text,
            type_l2 text,
            migoid_l2 text,
            l2_conn1 integer,
            l2_conn2 integer,
            reference_l2 text,
            comments text
        );

        create table ref_fiber_lease (
            lease_id text primary key,
            feature text,
            oid integer,
            start_range integer,
            end_range integer,
            reference text,
            lessee text,
            service_code integer,
            status_code integer,
            client text,
            reseau_code integer,
            ref_exploit text,
            comments text,
            migration_oid text
        );

        create table ref_isp_lease (
            lease_id text primary key,
            feature text,
            oid integer,
            reference text,
            lessee text,
            service_code integer,
            status_code integer,
            client text,
            reseau_code integer,
            ref_exploit text,
            comments text,
            migration_oid text,
            pair_oid integer,
            isp_container_id integer
        );

        create table ref_swag_interfaces (
            interface_id text primary key,
            hostname text,
            interface_name text,
            admin_status text,
            oper_status text,
            description text,
            link_type text,
            route_refs_json text,
            service_refs_json text,
            normalized_description text
        );

        create table ref_cpe_inventory (
            cpe_id text primary key,
            ip text,
            hostname text,
            vendor text,
            model text,
            version text,
            serial text,
            sfp text,
            sfp_model text,
            chassis text,
            normalized_hostname text
        );

        create table ref_cpe_configs (
            config_id text primary key,
            file_name text,
            hostname text,
            header_info text,
            service_refs_json text,
            route_refs_json text,
            vlans_json text,
            normalized_payload text
        );

        create table ref_network_devices (
            device_id text primary key,
            device_name text,
            source_file text,
            source_family text,
            vendor text,
            location_text text,
            announcement_text text,
            header_info text,
            client_hint text,
            site_hint text,
            route_refs_json text,
            service_refs_json text,
            vlans_json text,
            normalized_label text
        );

        create table ref_network_interfaces (
            network_interface_id text primary key,
            device_id text,
            device_name text,
            source_file text,
            source_family text,
            interface_name text,
            description text,
            service_refs_json text,
            route_refs_json text,
            vlan_ids_json text,
            normalized_label text
        );

        create table ref_network_vlans (
            network_vlan_id text primary key,
            device_id text,
            device_name text,
            source_file text,
            source_family text,
            vlan_id integer,
            label text,
            service_refs_json text,
            route_refs_json text,
            normalized_label text
        );

        create table party_master (
            party_id text primary key,
            canonical_name text,
            normalized_name text,
            party_type text,
            source_priority integer
        );

        create table party_alias (
            alias_id text primary key,
            party_id text,
            alias_value text,
            normalized_alias text,
            source_table text,
            source_key text
        );

        create table service_master_active (
            service_id text primary key,
            service_key text,
            nature_service text,
            principal_client text,
            principal_offer text,
            principal_external_ref text,
            principal_internal_ref text,
            route_refs_json text,
            service_refs_json text,
            endpoint_a_raw text,
            endpoint_z_raw text,
            client_final text,
            line_count integer,
            active_line_count integer
        );

        create table service_bss_line (
            service_id text,
            lea_line_id text,
            role_ligne text,
            is_principal integer
        );

        create table service_party (
            service_id text,
            role_name text,
            party_id text,
            rule_name text,
            score integer,
            source_table text,
            source_key text
        );

        create table service_endpoint (
            service_id text,
            endpoint_label text,
            raw_value text,
            matched_site_id text,
            matched_site_name text,
            score integer,
            rule_name text
        );

        create table service_support_optique (
            service_id text,
            route_ref text,
            route_id text,
            route_match_rule text,
            route_score integer,
            lease_ref text,
            lease_id text,
            lease_match_rule text,
            lease_score integer,
            fiber_lease_id text,
            fiber_lease_match_rule text,
            fiber_lease_score integer,
            isp_lease_id text,
            isp_lease_match_rule text,
            isp_lease_score integer,
            support_type text,
            support_ref text,
            logical_route_id text,
            cable_id text,
            cable_match_rule text,
            cable_score integer,
            housing_id text,
            housing_match_rule text,
            housing_score integer,
            site_a_optical_id text,
            site_z_optical_id text,
            optical_context_json text
        );

        create table service_support_reseau (
            service_id text,
            service_ref text,
            interface_id text,
            interface_match_rule text,
            interface_score integer,
            network_interface_id text,
            network_interface_match_rule text,
            network_interface_score integer,
            network_vlan_id text,
            network_vlan_match_rule text,
            network_vlan_score integer,
            cpe_id text,
            cpe_match_rule text,
            cpe_score integer,
            config_id text,
            config_match_rule text,
            config_score integer,
            inferred_vlans_json text
        );

        create table service_match_evidence (
            evidence_id text primary key,
            service_id text,
            evidence_type text,
            rule_name text,
            score integer,
            source_table text,
            source_key text,
            payload_json text
        );

        create table service_review_queue (
            review_id text primary key,
            service_id text,
            review_type text,
            severity text,
            current_state text,
            reason text,
            context_json text
        );

        create table gold_service_active (
            service_id text primary key,
            match_state text,
            confidence_band text,
            contract_party_id text,
            final_party_id text,
            endpoint_a_site_id text,
            endpoint_z_site_id text,
            route_ref text,
            route_id text,
            lease_id text,
            fiber_lease_id text,
            isp_lease_id text,
            interface_id text,
            network_interface_id text,
            network_vlan_id text,
            cpe_id text,
            config_id text,
            inferred_vlans_json text,
            strong_evidence_count integer,
            evidence_count integer,
            summary_json text
        );

        create table service_facturable_final (
            service_id text primary key,
            service_key text,
            principal_lea_line_id text,
            lea_line_ids_json text,
            line_count integer,
            active_line_count integer,
            nature_service text,
            principal_offer text,
            principal_external_ref text,
            principal_internal_ref text,
            principal_client_raw text,
            client_final_raw text,
            contract_party_id text,
            contract_party_name text,
            final_party_id text,
            final_party_name text,
            final_party_source text,
            endpoint_a_raw text,
            endpoint_z_raw text,
            site_a_id text,
            site_a_name text,
            site_a_source text,
            site_z_id text,
            site_z_name text,
            site_z_source text,
            route_ref text,
            route_id text,
            lease_id text,
            fiber_lease_id text,
            isp_lease_id text,
            logical_route_id text,
            cable_id text,
            housing_id text,
            optical_site_a_id text,
            optical_site_z_id text,
            optical_context_json text,
            optical_source text,
            agent_optical_support_hint text,
            interface_id text,
            network_interface_id text,
            network_vlan_id text,
            cpe_id text,
            config_id text,
            inferred_vlans_json text,
            network_source text,
            agent_network_support_hint text,
            gold_match_state text,
            gold_confidence_band text,
            agent_resolution_status text,
            agent_resolution_confidence text,
            publication_status text,
            selected_truth_source text,
            gap_flags_json text,
            publication_comment text,
            published_at text
        );

        create table override_party_alias (
            normalized_alias text primary key,
            forced_party_name text,
            comment text
        );

        create table override_site_alias (
            normalized_alias text primary key,
            forced_site_id text,
            comment text
        );

        create table override_service_match (
            service_id text,
            match_type text,
            forced_target_id text,
            comment text
        );

        create index idx_lea_grouping_key on lea_active_lines(grouping_key);
        create index idx_ref_sites_norm_ref on ref_sites(normalized_reference);
        create index idx_ref_sites_norm_addr on ref_sites(normalized_address);
        create index idx_ref_routes_ref on ref_routes(route_ref);
        create index idx_ref_lease_template_ref on ref_lease_template(ref_exploit);
        create index idx_ref_fiber_lease_ref on ref_fiber_lease(ref_exploit);
        create index idx_ref_isp_lease_ref on ref_isp_lease(ref_exploit);
        create index idx_ref_swag_service on ref_swag_interfaces(service_refs_json);
        create index idx_ref_network_interface_device on ref_network_interfaces(device_name);
        create index idx_ref_network_vlan_device on ref_network_vlans(device_name, vlan_id);
        create index idx_service_match_service on service_match_evidence(service_id, score);
        """
    )
    con.commit()


def load_lea_active(con: sqlite3.Connection) -> None:
    LOG.info("Loading active LEA lines")
    xl = _require_dependency(openpyxl, "openpyxl")
    wb = xl.load_workbook(LEA_PATH, read_only=True, data_only=True)
    ws = wb["GLOBAL"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    index = {header: idx for idx, header in enumerate(headers)}

    def value(row: tuple[object, ...], name: str) -> object:
        return row[index[name]] if name in index else None

    records = []
    for source_row, row in enumerate(rows, start=2):
        status_code = str(value(row, "CMD - Statut Commande") or "").strip()
        if status_code != "40":
            continue
        command_internal = str(value(row, "CMD - Numéro commande interne") or "")
        contract_file = str(value(row, "nom fichier") or "")
        is_old = 1 if "OLD" in norm_text(command_internal + " " + contract_file) else 0
        offer_label = str(value(row, "CMD.OFF - Libellé détaillé du code offre") or "")
        nature_service, role_ligne = classify_offer(offer_label)
        command_external = str(value(row, "CMD - Numéro commande externe") or "")
        endpoint_a_raw = str(value(row, "CMD - Secteur géographique1") or "")
        endpoint_z_raw = str(value(row, "CMD - Secteur géographique2") or "")
        client_contractant = str(value(row, "CMD - Nom client contractant") or "")
        client_final = str(value(row, "Client Final (ADV)") or "")
        route_refs = extract_route_refs(command_external, contract_file)
        service_refs = extract_service_refs(command_external, contract_file)
        grouping_key = build_grouping_key(
            client_contractant,
            nature_service,
            role_ligne,
            route_refs,
            service_refs,
            endpoint_a_raw,
            endpoint_z_raw,
            client_final,
            contract_file,
            command_internal,
        )
        lea_line_id = f"LEA-{safe_hash([client_contractant, command_internal, offer_label, source_row])}"
        records.append(
            (
                lea_line_id,
                client_contractant,
                offer_label,
                nature_service,
                role_ligne,
                command_internal,
                command_external,
                contract_file,
                endpoint_a_raw,
                endpoint_z_raw,
                client_final,
                str(value(row, "Linéaire installation LigneDeCmd") or ""),
                str(value(row, "CMD - Date de signature") or ""),
                str(value(row, "Date création LigneDeCmd") or ""),
                str(value(row, "Durée Service LigneDeCmd") or ""),
                status_code,
                str(value(row, "CMD - Date de livraison ADV") or ""),
                str(value(row, "CMD - Date de résiliation") or ""),
                str(value(row, "CMD - FMS") or ""),
                str(value(row, "RM - Initiale") or ""),
                str(value(row, "RM - Dernière") or ""),
                str(value(row, "CMD - IRU") or ""),
                str(value(row, "Montant LigneDeCmd") or ""),
                json.dumps(route_refs, ensure_ascii=True),
                json.dumps(service_refs, ensure_ascii=True),
                is_old,
                grouping_key,
                source_row,
            )
        )

    con.executemany(
        """
        insert into lea_active_lines values (
            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?
        )
        """,
        records,
    )
    con.commit()
    LOG.info("Loaded %s active LEA lines", len(records))


def build_grouping_key(
    client_contractant: str,
    nature_service: str,
    role_ligne: str,
    route_refs: list[str],
    service_refs: list[str],
    endpoint_a_raw: str,
    endpoint_z_raw: str,
    client_final: str,
    contract_file: str,
    command_internal: str,
) -> str:
    principal_marker = norm_text(client_contractant)
    if service_refs:
        return f"SVCREF|{service_refs[0]}|{principal_marker}|{nature_service}"
    if route_refs:
        return f"ROUTE|{route_refs[0]}|{principal_marker}|{nature_service}"
    if norm_text(contract_file):
        return f"CONTRACT|{safe_hash([client_contractant, contract_file, nature_service])}"
    if role_ligne != "principal":
        return f"AUX|{safe_hash([client_contractant, nature_service, endpoint_a_raw, endpoint_z_raw, client_final])}"
    return f"FALLBACK|{safe_hash([client_contractant, nature_service, endpoint_a_raw, endpoint_z_raw, client_final, command_internal])}"


def iter_gdb_records(layer: str) -> Iterator[dict[str, object]]:
    fiona_mod = _require_dependency(fiona, "fiona")
    with fiona_mod.open(GDB_URI, layer=layer) as src:
        for feature in src:
            yield dict(feature["properties"])


def load_sites(con: sqlite3.Connection) -> None:
    LOG.info("Loading GDB sites")
    records = []
    for item in iter_gdb_records("Hubsite"):
        records.append(
            (
                str(item.get("MIGRATION_OID") or ""),
                str(item.get("MIGRATION_ID") or ""),
                str(item.get("REFERENCE") or ""),
                str(item.get("USERREFERENCE") or ""),
                str(item.get("ADRESSE1") or ""),
                item.get("FONCTION_DU_SITE"),
                item.get("RESEAU"),
                item.get("GESTIONNAIRE"),
                item.get("PROPRIETAIRE"),
                item.get("PRECISION"),
                str(item.get("CODE_PROJET") or ""),
                norm_text(item.get("REFERENCE")),
                norm_text(item.get("USERREFERENCE")),
                norm_text(item.get("ADRESSE1")),
            )
        )
    con.executemany("insert into ref_sites values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", records)
    con.commit()
    LOG.info("Loaded %s sites", len(records))


def _hubsite_lookup_by_oid(con: sqlite3.Connection) -> dict[str, tuple[str, str]]:
    return {
        row[0]: (row[0], row[2] or row[0])
        for row in con.execute("select site_id, migration_id, reference from ref_sites")
        if row[0]
    }


def _site_name_by_id_map(con: sqlite3.Connection) -> dict[str, str]:
    return {
        row[0]: row[2] or row[0]
        for row in con.execute("select site_id, migration_id, reference from ref_sites")
        if row[0]
    }


def _optical_site_link_map(con: sqlite3.Connection) -> dict[tuple[str, str], tuple[str, str, str, int]]:
    links: dict[tuple[str, str], tuple[str, str, str, int]] = {}
    for row in con.execute(
        "select object_type, object_migration_oid, site_id, site_name, rule_name, score from ref_optical_site_link"
    ):
        key = (row[0], row[1])
        current = links.get(key)
        if current is None or (row[5] or 0) > current[3]:
            links[key] = (row[2], row[3], row[4], row[5] or 0)
    return links


def _site_match_from_values(
    values: Iterable[object],
    site_index: dict[str, list[tuple]],
    all_sites: list[tuple],
) -> SiteMatch:
    best = SiteMatch(None, 0, None, None)
    for value in values:
        match = match_site(str(value or ""), site_index, all_sites)
        if match.score > best.score:
            best = match
    return best


def load_routes(con: sqlite3.Connection) -> None:
    LOG.info("Loading GDB optical topology")
    site_index, all_sites = build_site_index(con)
    hubsite_lookup = _hubsite_lookup_by_oid(con)

    parent_by_child: dict[str, tuple[str, str | None]] = {}
    for item in iter_gdb_records("IMPORT_ISP_TEMPLATE"):
        child_oid = str(item.get("ISPMIGOID") or "").strip()
        if not child_oid:
            continue
        parent_by_child[child_oid] = (
            str(item.get("PARENTTYPE") or "").strip(),
            str(item.get("PARENTMIGOID") or "").strip() or None,
        )

    site_link_records: list[tuple] = []
    for layer, (object_type, object_oid_field, hubsite_oid_field) in HOUSING_SITE_RELATION_LAYERS.items():
        for item in iter_gdb_records(layer):
            object_oid = str(item.get(object_oid_field) or "").strip()
            hubsite_oid = str(item.get(hubsite_oid_field) or "").strip()
            if not object_oid or not hubsite_oid or hubsite_oid not in hubsite_lookup:
                continue
            site_id, site_name = hubsite_lookup[hubsite_oid]
            site_link_records.append(
                (object_type, object_oid, site_id, site_name, "gdb_relation", 100, layer)
            )

    housing_records: list[tuple] = []
    for layer in ("Rack", "OptPatchPanel", "Room", "Chamber", "Enclosure", "Pedestal", "ISPEnclosure", "ISPManifold"):
        for item in iter_gdb_records(layer):
            migration_oid = str(item.get("MIGRATION_OID") or "").strip()
            if not migration_oid:
                continue
            reference = str(item.get("REFERENCE") or "")
            userreference = str(item.get("USERREFERENCE") or "")
            description = str(item.get("DESCRIPTION") or "")
            comments = str(item.get("COMMENTS") or "")
            location = str(item.get("LOCATION") or "")
            site_match = _site_match_from_values(
                (reference, userreference, description, comments, location),
                site_index,
                all_sites,
            )
            parent_type, parent_migration_oid = parent_by_child.get(migration_oid, (None, None))
            housing_id = f"HSG-{layer.upper()}-{safe_hash([migration_oid, reference, userreference])}"
            housing_records.append(
                (
                    housing_id,
                    migration_oid,
                    layer,
                    reference,
                    userreference,
                    description,
                    comments,
                    location,
                    parent_type,
                    parent_migration_oid,
                    site_match.site_id,
                    site_match.site_name,
                    site_match.rule,
                    site_match.score,
                )
            )

    connection_records: list[tuple] = []
    for idx, item in enumerate(iter_gdb_records("CONNEXION_TEMPLATE"), start=1):
        connection_records.append(
            (
                f"OCN-{idx:07d}-{safe_hash([item.get('HOUSINGMIGOID'), item.get('OBJ1_MIGOID'), item.get('OBJ2_MIGOID'), item.get('OBJ1_CONNECTOR1'), item.get('OBJ2_CONNECTOR1')])}",
                str(item.get("HOUSING_TYPE") or ""),
                str(item.get("HOUSINGMIGOID") or ""),
                str(item.get("OBJ1_TYPE") or ""),
                str(item.get("OBJ1_MIGOID") or ""),
                item.get("OBJ1_CONNECTOR1"),
                item.get("OBJ1_CONNECTOR2"),
                str(item.get("OBJ2_TYPE") or ""),
                str(item.get("OBJ2_MIGOID") or ""),
                item.get("OBJ2_CONNECTOR1"),
                item.get("OBJ2_CONNECTOR2"),
                str(item.get("TYPE_BRANCHEMENT") or ""),
                str(item.get("TRAY_MIGRATIONID") or ""),
                item.get("ID_RATTACHEMENT"),
            )
        )

    cable_records: list[tuple] = []
    cable_hint_records: list[tuple] = []
    for item in iter_gdb_records("Fiber_Cable"):
        migration_oid = str(item.get("MIGRATION_OID") or "").strip()
        if not migration_oid:
            continue
        reference = str(item.get("REFERENCE") or "")
        userreference = str(item.get("USERREFERENCE") or "")
        comments = str(item.get("COMMENTS") or "")
        labeltext = str(item.get("LABELTEXT") or "")
        location = str(item.get("LOCATION") or "")
        cable_id = f"CAB-{safe_hash([migration_oid, reference, userreference])}"
        site_tokens = extract_place_tokens(reference, userreference, comments)
        fiber_count = extract_fiber_count(reference, userreference, comments, labeltext)
        cable_records.append(
            (
                cable_id,
                migration_oid,
                str(item.get("MIGRATION_ID") or ""),
                reference,
                userreference,
                comments,
                labeltext,
                location,
                fiber_count,
                str(item.get("CABLETYPE") or ""),
                str(item.get("RESEAU") or ""),
                str(item.get("CODE_PROJET") or ""),
                str(item.get("STATUS") or ""),
                norm_text(reference or userreference or comments),
                json.dumps(site_tokens, ensure_ascii=True),
            )
        )
        for token in site_tokens:
            base_score = 90 if fiber_count is not None and fiber_count <= 24 else 75
            cable_hint_records.append((cable_id, token, base_score, "fiber_cable_name_token"))

    con.executemany(
        "insert into ref_optical_housing values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        housing_records,
    )
    con.executemany(
        "insert into ref_optical_connection values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        connection_records,
    )
    con.executemany(
        "insert into ref_optical_site_link values (?,?,?,?,?,?,?)",
        site_link_records,
    )
    con.executemany(
        "insert into ref_optical_cable values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        cable_records,
    )
    con.executemany(
        "insert into ref_optical_cable_site_hint values (?,?,?,?)",
        cable_hint_records,
    )
    con.commit()
    LOG.info(
        "Loaded %s housings, %s connections, %s cable records, %s direct site links",
        len(housing_records),
        len(connection_records),
        len(cable_records),
        len(site_link_records),
    )


def load_lease_tables(con: sqlite3.Connection) -> None:
    LOG.info("Loading GDB lease tables")
    site_index, all_sites = build_site_index(con)
    housing_site_links = _optical_site_link_map(con)
    site_name_by_id = _site_name_by_id_map(con)

    route_rows_by_ref: dict[str, tuple[str, str, str, str, str, str]] = {}
    parcours_rows: list[tuple] = []
    logical_route_records: list[tuple] = []
    optical_lease_records: list[tuple] = []
    optical_lease_endpoint_records: list[tuple] = []

    def best_site_for_endpoint(
        housing_type: object,
        housing_migoid: object,
        object_type: object,
        object_migoid: object,
        reference_label: object,
    ) -> SiteMatch:
        candidates = [
            (str(object_type or ""), str(object_migoid or "").strip()),
            (str(housing_type or ""), str(housing_migoid or "").strip()),
        ]
        for object_kind, migration_oid in candidates:
            if migration_oid and (object_kind, migration_oid) in housing_site_links:
                site_id, site_name, rule_name, score = housing_site_links[(object_kind, migration_oid)]
                return SiteMatch(site_id, score, rule_name, site_name)
        return _site_match_from_values((reference_label,), site_index, all_sites)

    def add_logical_route(
        route_ref: str | None,
        *,
        source_layer: str,
        source_oid: object,
        reference: object = None,
        path_id: object = None,
        pair_oid: object = None,
        feature_type: object = None,
        lessee: object = None,
        client: object = None,
        network: object = None,
        status: object = None,
        comments: object = None,
    ) -> str | None:
        route_ref_value = str(route_ref or "").strip().upper() or None
        reference_value = str(reference or "").strip() or None
        if not route_ref_value and not reference_value:
            return None
        logical_route_id = f"OROUTE-{safe_hash([source_layer, source_oid, route_ref_value or '', reference_value or '', path_id or '', pair_oid or ''])}"
        logical_route_records.append(
            (
                logical_route_id,
                route_ref_value,
                source_layer,
                str(source_oid or ""),
                route_ref_value,
                reference_value,
                str(path_id or ""),
                str(pair_oid or ""),
                str(feature_type or ""),
                str(lessee or ""),
                str(client or ""),
                str(network or ""),
                str(status or ""),
                str(comments or ""),
                norm_text(route_ref_value or reference_value or comments),
            )
        )
        if route_ref_value and route_ref_value not in route_rows_by_ref:
            route_rows_by_ref[route_ref_value] = (
                logical_route_id,
                route_ref_value,
                str(network or ""),
                str(client or ""),
                str(lessee or ""),
                str(status or ""),
            )
        return logical_route_id

    lease_template_records = []
    for idx, item in enumerate(iter_gdb_records("LEASE_TEMPLATE"), start=1):
        lease_id = f"LEASET-{idx:06d}-{safe_hash([item.get('REF_EXPLOIT'), item.get('MIGOIDL1'), item.get('MIGOIDL2'), item.get('L1_CONN1'), item.get('L2_CONN1')])}"
        route_ref = str(item.get("REF_EXPLOIT") or "").strip().upper()
        lease_template_records.append(
            (
                lease_id,
                route_ref,
                str(item.get("RESEAU") or ""),
                str(item.get("LESSEE") or ""),
                str(item.get("CLIENTS") or ""),
                str(item.get("HOUSINGTYPEL1") or ""),
                str(item.get("HOUSINGMIGOIDL1") or ""),
                str(item.get("TYPEL1") or ""),
                str(item.get("MIGOIDL1") or ""),
                item.get("L1_CONN1"),
                item.get("L1_CONN2"),
                str(item.get("REFERENCEL1") or ""),
                str(item.get("HOUSINGTYPEL2") or ""),
                str(item.get("HOUSINGMIGOIDL2") or ""),
                str(item.get("TYPEL2") or ""),
                str(item.get("MIGOIDL2") or ""),
                item.get("L2_CONN1"),
                item.get("L2_CONN2"),
                str(item.get("REFERENCEL2") or ""),
                str(item.get("COMMENTS") or ""),
            )
        )
        optical_lease_records.append(
            (
                lease_id,
                "template",
                "LEASE_TEMPLATE",
                lease_id,
                route_ref,
                None,
                "",
                "",
                "",
                str(item.get("LESSEE") or ""),
                str(item.get("CLIENTS") or ""),
                str(item.get("RESEAU") or ""),
                "",
                str(item.get("COMMENTS") or ""),
                norm_text(route_ref or item.get("COMMENTS")),
            )
        )
        logical_route_id = add_logical_route(
            route_ref,
            source_layer="LEASE_TEMPLATE",
            source_oid=lease_id,
            reference=None,
            lessee=item.get("LESSEE"),
            client=item.get("CLIENTS"),
            network=item.get("RESEAU"),
            comments=item.get("COMMENTS"),
        )
        for endpoint_label, housing_type, housing_migoid, object_type, object_migoid, conn1, conn2, reference_label, step_no, step_type in (
            ("L1", item.get("HOUSINGTYPEL1"), item.get("HOUSINGMIGOIDL1"), item.get("TYPEL1"), item.get("MIGOIDL1"), item.get("L1_CONN1"), item.get("L1_CONN2"), item.get("REFERENCEL1"), 1, "origin"),
            ("L2", item.get("HOUSINGTYPEL2"), item.get("HOUSINGMIGOIDL2"), item.get("TYPEL2"), item.get("MIGOIDL2"), item.get("L2_CONN1"), item.get("L2_CONN2"), item.get("REFERENCEL2"), 2, "destination"),
        ):
            site_match = best_site_for_endpoint(housing_type, housing_migoid, object_type, object_migoid, reference_label)
            optical_lease_endpoint_records.append(
                (
                    lease_id,
                    endpoint_label,
                    str(housing_type or ""),
                    str(housing_migoid or ""),
                    str(object_type or ""),
                    str(object_migoid or ""),
                    conn1,
                    conn2,
                    str(reference_label or ""),
                    site_match.site_id,
                    site_match.site_name,
                    site_match.rule,
                    site_match.score,
                )
            )
            if logical_route_id:
                parcours_rows.append(
                    (
                        route_ref,
                        logical_route_id,
                        step_no,
                        step_type,
                        site_match.site_name or str(reference_label or ""),
                        str(housing_type or ""),
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        str(reference_label or ""),
                    )
                )
    con.executemany("insert into ref_lease_template values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", lease_template_records)

    fiber_lease_records = []
    for idx, item in enumerate(iter_gdb_records("Fiber_Lease"), start=1):
        base_lease_id = str(item.get("MIGRATION_OID") or f"FLEASE-{item.get('OID')}-{item.get('STARTRANGE')}-{item.get('ENDRANGE')}")
        lease_id = f"{base_lease_id}-{idx:06d}"
        fiber_lease_records.append(
            (
                lease_id,
                str(item.get("FEATURE") or ""),
                item.get("OID"),
                item.get("STARTRANGE"),
                item.get("ENDRANGE"),
                str(item.get("REFERENCE") or ""),
                str(item.get("LESSEE") or ""),
                item.get("SERVICE"),
                item.get("STATUS"),
                str(item.get("CLIENT") or ""),
                item.get("RESEAU"),
                str(item.get("REF_EXPLOIT") or "").strip().upper(),
                str(item.get("COMMENTS") or ""),
                str(item.get("MIGRATION_OID") or ""),
            )
        )
        route_ref = str(item.get("REF_EXPLOIT") or "").strip().upper() or infer_route_ref(item.get("REFERENCE"), item.get("COMMENTS"))
        optical_lease_records.append(
            (
                lease_id,
                "fiber_lease",
                "Fiber_Lease",
                str(item.get("MIGRATION_OID") or item.get("OID") or lease_id),
                route_ref or "",
                str(item.get("REFERENCE") or ""),
                str(item.get("PATHID") or ""),
                str(item.get("PAIROID") or ""),
                str(item.get("FEATURE") or ""),
                str(item.get("LESSEE") or ""),
                str(item.get("CLIENT") or ""),
                str(item.get("RESEAU") or ""),
                str(item.get("STATUS") or ""),
                str(item.get("COMMENTS") or ""),
                norm_text(route_ref or item.get("REFERENCE") or item.get("COMMENTS")),
            )
        )
        add_logical_route(
            route_ref,
            source_layer="Fiber_Lease",
            source_oid=item.get("MIGRATION_OID") or item.get("OID") or lease_id,
            reference=item.get("REFERENCE"),
            path_id=item.get("PATHID"),
            pair_oid=item.get("PAIROID"),
            feature_type=item.get("FEATURE"),
            lessee=item.get("LESSEE"),
            client=item.get("CLIENT"),
            network=item.get("RESEAU"),
            status=item.get("STATUS"),
            comments=item.get("COMMENTS"),
        )
    con.executemany("insert into ref_fiber_lease values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", fiber_lease_records)

    isp_lease_records = []
    for idx, item in enumerate(iter_gdb_records("ISPLease"), start=1):
        base_lease_id = str(item.get("MIGRATION_OID") or f"ISPLEASE-{item.get('OID')}-{item.get('ISPCONTAINERID')}")
        lease_id = f"{base_lease_id}-{idx:06d}"
        isp_lease_records.append(
            (
                lease_id,
                str(item.get("FEATURE") or ""),
                item.get("OID"),
                str(item.get("REFERENCE") or ""),
                str(item.get("LESSEE") or ""),
                item.get("SERVICE"),
                item.get("STATUS"),
                str(item.get("CLIENT") or ""),
                item.get("RESEAU"),
                str(item.get("REF_EXPLOIT") or "").strip().upper(),
                str(item.get("COMMENTS") or ""),
                str(item.get("MIGRATION_OID") or ""),
                item.get("PAIROID"),
                item.get("ISPCONTAINERID"),
            )
        )
        route_ref = str(item.get("REF_EXPLOIT") or "").strip().upper() or infer_route_ref(item.get("REFERENCE"), item.get("COMMENTS"))
        optical_lease_records.append(
            (
                lease_id,
                "isp_lease",
                "ISPLease",
                str(item.get("MIGRATION_OID") or item.get("OID") or lease_id),
                route_ref or "",
                str(item.get("REFERENCE") or ""),
                str(item.get("PATHID") or ""),
                str(item.get("PAIROID") or ""),
                str(item.get("FEATURE") or ""),
                str(item.get("LESSEE") or ""),
                str(item.get("CLIENT") or ""),
                str(item.get("RESEAU") or ""),
                str(item.get("STATUS") or ""),
                str(item.get("COMMENTS") or ""),
                norm_text(route_ref or item.get("REFERENCE") or item.get("COMMENTS")),
            )
        )
        add_logical_route(
            route_ref,
            source_layer="ISPLease",
            source_oid=item.get("MIGRATION_OID") or item.get("OID") or lease_id,
            reference=item.get("REFERENCE"),
            path_id=item.get("PATHID"),
            pair_oid=item.get("PAIROID"),
            feature_type=item.get("FEATURE"),
            lessee=item.get("LESSEE"),
            client=item.get("CLIENT"),
            network=item.get("RESEAU"),
            status=item.get("STATUS"),
            comments=item.get("COMMENTS"),
        )
    con.executemany("insert into ref_isp_lease values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", isp_lease_records)
    con.executemany(
        "insert into ref_optical_lease values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        optical_lease_records,
    )
    con.executemany(
        "insert into ref_optical_lease_endpoint values (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        optical_lease_endpoint_records,
    )
    logical_route_rows = []
    seen_logical_route_ids: set[str] = set()
    for row in logical_route_records:
        if row[0] in seen_logical_route_ids:
            continue
        logical_route_rows.append(row)
        seen_logical_route_ids.add(row[0])
    con.executemany(
        "insert into ref_optical_logical_route values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        logical_route_rows,
    )
    con.executemany(
        "insert into ref_routes values (?,?,?,?,?,?)",
        list(route_rows_by_ref.values()),
    )
    con.executemany(
        "insert into ref_route_parcours values (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        parcours_rows,
    )
    con.commit()
    LOG.info(
        "Loaded %s lease_template, %s fiber_lease, %s isp_lease rows, %s logical routes",
        len(lease_template_records),
        len(fiber_lease_records),
        len(isp_lease_records),
        len(logical_route_rows),
    )


def load_swag_interfaces(con: sqlite3.Connection) -> None:
    LOG.info("Loading SWAG interfaces")
    xl = _require_dependency(openpyxl, "openpyxl")
    wb = xl.load_workbook(SWAG_PATH, read_only=True, data_only=True)
    ws = wb["inventaire interface"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    records = []
    for row in rows:
        hostname = str(row[0] or "")
        interface_name = str(row[1] or "")
        description = str(row[5] or "")
        interface_id = f"IF-{safe_hash([hostname, interface_name])}"
        records.append(
            (
                interface_id,
                hostname,
                interface_name,
                str(row[3] or ""),
                str(row[4] or ""),
                description,
                str(row[6] or ""),
                json.dumps(extract_route_refs(description), ensure_ascii=True),
                json.dumps(extract_service_refs(description), ensure_ascii=True),
                norm_text(description),
            )
        )
    con.executemany("insert into ref_swag_interfaces values (?,?,?,?,?,?,?,?,?,?)", records)
    con.commit()
    LOG.info("Loaded %s SWAG interfaces", len(records))


def load_cpe_inventory(con: sqlite3.Connection) -> None:
    LOG.info("Loading CPE inventory")
    xl = _require_dependency(openpyxl, "openpyxl")
    wb = xl.load_workbook(CPE_PATH, read_only=True, data_only=True)
    ws = wb["Audit_Inventaire_2901"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    records = []
    for row in rows:
        hostname = str(row[2] or "")
        if not hostname:
            continue
        cpe_id = f"CPE-{safe_hash([hostname, row[0]])}"
        records.append(
            (
                cpe_id,
                str(row[0] or ""),
                hostname,
                str(row[3] or ""),
                str(row[4] or ""),
                str(row[5] or ""),
                str(row[6] or ""),
                str(row[8] or ""),
                str(row[9] or ""),
                str(row[10] or ""),
                norm_text(hostname),
            )
        )
    con.executemany("insert into ref_cpe_inventory values (?,?,?,?,?,?,?,?,?,?,?)", records)
    con.commit()
    LOG.info("Loaded %s CPE inventory rows", len(records))


def load_cpe_configs(con: sqlite3.Connection) -> None:
    LOG.info("Loading CPE configs")
    records = []
    for path in sorted(CONFIG_DIR.glob("*.txt")):
        payload = path.read_text(encoding="latin1", errors="ignore")
        hostname = ""
        match = HOSTNAME_PATTERN.search(payload)
        if match:
            hostname = next(group for group in match.groups() if group)
        header_match = HEADER_PATTERN.search(payload)
        header_info = header_match.group(1) if header_match else ""
        service_refs = extract_service_refs(payload)
        route_refs = extract_route_refs(payload)
        vlans = sorted({int(vlan) for vlan in VLAN_LINE_PATTERN.findall(payload)})
        config_id = f"CFG-{safe_hash([path.name, hostname])}"
        records.append(
            (
                config_id,
                path.name,
                hostname,
                header_info,
                json.dumps(service_refs, ensure_ascii=True),
                json.dumps(route_refs, ensure_ascii=True),
                json.dumps(vlans, ensure_ascii=True),
                norm_text(payload[:2000]),
            )
        )
    con.executemany("insert into ref_cpe_configs values (?,?,?,?,?,?,?,?)", records)
    con.commit()
    LOG.info("Loaded %s config files", len(records))


def build_service_master(con: sqlite3.Connection) -> None:
    LOG.info("Building service master")
    rows = con.execute(
        """
        select lea_line_id, client_contractant, offer_label, nature_service, role_ligne,
               command_internal, command_external, endpoint_a_raw, endpoint_z_raw,
               client_final, route_refs_json, service_refs_json, grouping_key, is_old
        from lea_active_lines
        where is_old = 0
        order by client_contractant, offer_label, lea_line_id
        """
    ).fetchall()

    groups: dict[str, list[sqlite3.Row | tuple]] = defaultdict(list)
    for row in rows:
        groups[row[12]].append(row)

    service_master_records = []
    service_bss_records = []
    counter = 1
    for service_key, members in groups.items():
        principal = next((row for row in members if row[4] == "principal"), members[0])
        service_id = f"SVC-{counter:05d}"
        counter += 1
        route_refs = merge_json_lists(row[10] for row in members)
        service_refs = merge_json_lists(row[11] for row in members)
        service_master_records.append(
            (
                service_id,
                service_key,
                principal[3],
                principal[1],
                principal[2],
                principal[6],
                principal[5],
                json.dumps(route_refs, ensure_ascii=True),
                json.dumps(service_refs, ensure_ascii=True),
                principal[7],
                principal[8],
                principal[9],
                len(members),
                len(members),
            )
        )
        for member in members:
            service_bss_records.append((service_id, member[0], member[4], 1 if member[0] == principal[0] else 0))

    con.executemany("insert into service_master_active values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", service_master_records)
    con.executemany("insert into service_bss_line values (?,?,?,?)", service_bss_records)
    con.commit()
    LOG.info("Built %s active services", len(service_master_records))


def merge_json_lists(values: Iterable[str]) -> list[str]:
    merged: list[str] = []
    seen = set()
    for value in values:
        for item in json.loads(value or "[]"):
            if item not in seen:
                merged.append(item)
                seen.add(item)
    return merged


def build_site_index(con: sqlite3.Connection) -> tuple[dict[str, list[tuple]], list[tuple]]:
    rows = con.execute(
        "select site_id, migration_id, reference, userreference, address1, normalized_reference, normalized_userreference, normalized_address from ref_sites"
    ).fetchall()
    alias_index: dict[str, list[tuple]] = defaultdict(list)
    for row in rows:
        site_id, migration_id, reference, userreference, address1, norm_ref, norm_user, norm_address = row
        for alias in [site_id, migration_id, reference, userreference, address1, norm_ref, norm_user, norm_address]:
            alias_key = norm_text(alias)
            if alias_key:
                alias_index[alias_key].append(row)
    return alias_index, rows


def score_label_match(seeds: Iterable[str], candidate: str) -> int:
    candidate_clean = clean_business_label(candidate)
    if not candidate_clean:
        return 0
    candidate_tokens = business_tokens(candidate_clean)
    best_score = 0
    for seed in seeds:
        seed_clean = clean_business_label(seed)
        if not seed_clean:
            continue
        if seed_clean == candidate_clean:
            best_score = max(best_score, 96)
            continue
        if seed_clean in candidate_clean or candidate_clean in seed_clean:
            overlap = len(business_tokens(seed_clean) & candidate_tokens)
            if overlap >= 2:
                best_score = max(best_score, 90)
                continue
        overlap = len(business_tokens(seed_clean) & candidate_tokens)
        if overlap >= 3:
            best_score = max(best_score, min(88, 18 * overlap + 20))
        elif overlap == 2:
            best_score = max(best_score, 72)
    return best_score


def best_label_candidate(seeds: Iterable[str], candidates: list[tuple], label_index: int) -> tuple | None:
    best_row = None
    best_score = 0
    second_score = 0
    for row in candidates:
        score = score_label_match(seeds, row[label_index])
        if score > best_score:
            second_score = best_score
            best_score = score
            best_row = row
        elif score > second_score:
            second_score = score
    if best_row and best_score >= 72 and (best_score - second_score >= 10 or best_score >= 90):
        return (*best_row, best_score)
    return None


def party_id_for_alias(alias_index: dict[str, str], value: str) -> str | None:
    normalized = norm_text(value)
    if normalized in alias_index:
        return alias_index[normalized]
    cleaned = norm_text(clean_business_label(value))
    if cleaned in alias_index:
        return alias_index[cleaned]
    return None


def match_site(raw_value: str, alias_index: dict[str, list[tuple]], all_sites: list[tuple]) -> SiteMatch:
    normalized = norm_text(raw_value)
    if not normalized:
        return SiteMatch(None, 0, None, None)
    if normalized in alias_index:
        row = alias_index[normalized][0]
        return SiteMatch(row[0], 100, "site_exact", row[2] or row[3])

    best_row = None
    best_score = 0
    tokens = {token for token in normalized.split() if len(token) >= 4}
    for row in all_sites:
        site_tokens = set((row[5] or "").split()) | set((row[6] or "").split()) | set((row[7] or "").split())
        overlap = len(tokens & site_tokens)
        if overlap >= 2:
            score = min(85, overlap * 20)
            if score > best_score:
                best_score = score
                best_row = row
    if best_row:
        return SiteMatch(best_row[0], best_score, "site_token_overlap", best_row[2] or best_row[3])
    return SiteMatch(None, 0, None, None)


def reconcile_services(con: sqlite3.Connection) -> None:
    LOG.info("Reconciling service endpoints and OSS supports")
    site_index, all_sites = build_site_index(con)

    party_alias_index = {row[0]: row[1] for row in con.execute("select normalized_alias, party_id from party_alias")}
    cpe_by_hostname = {norm_text(row[1]): (row[0], row[1]) for row in con.execute("select cpe_id, hostname from ref_cpe_inventory")}

    route_index = {
        (row[1] or "").strip().upper(): row
        for row in con.execute("select route_id, route_ref, network, client, lessee, status from ref_routes")
        if row[1]
    }
    logical_route_by_ref = defaultdict(list)
    for row in con.execute(
        """
        select logical_route_id, route_ref, source_layer, source_oid, reference, path_id,
               pair_oid, feature_type, lessee, client, network, status, comments
        from ref_optical_logical_route
        """
    ):
        route_ref = (row[1] or "").strip().upper()
        if route_ref:
            logical_route_by_ref[route_ref].append(row)

    lease_rows = list(
        con.execute(
            "select lease_id, ref_exploit, reseau, lessee, clients, reference_l1, reference_l2, comments from ref_lease_template"
        )
    )
    lease_by_ref = defaultdict(list)
    lease_pair_index = defaultdict(list)
    for row in lease_rows:
        ref_exploit = (row[1] or "").strip().upper()
        if ref_exploit:
            lease_by_ref[ref_exploit].append(row)
    lease_endpoint_rows = defaultdict(dict)
    for row in con.execute(
        """
        select optical_lease_id, endpoint_label, site_id, site_name, site_score
        from ref_optical_lease_endpoint
        """
    ):
        lease_endpoint_rows[row[0]][row[1]] = row
    for row in lease_rows:
        lease_id = row[0]
        endpoint_map = lease_endpoint_rows.get(lease_id, {})
        site_l1_row = endpoint_map.get("L1")
        site_l2_row = endpoint_map.get("L2")
        if site_l1_row and site_l2_row and site_l1_row["site_id"] and site_l2_row["site_id"]:
            site_l1 = SiteMatch(
                site_l1_row["site_id"],
                site_l1_row["site_score"] or 0,
                "lease_endpoint_site",
                site_l1_row["site_name"],
            )
            site_l2 = SiteMatch(
                site_l2_row["site_id"],
                site_l2_row["site_score"] or 0,
                "lease_endpoint_site",
                site_l2_row["site_name"],
            )
        else:
            site_l1 = match_site(row[5] or "", site_index, all_sites)
            site_l2 = match_site(row[6] or "", site_index, all_sites)
        if site_l1.site_id and site_l2.site_id:
            lease_pair_index[tuple(sorted([site_l1.site_id, site_l2.site_id]))].append((row, site_l1, site_l2))

    fiber_by_ref = defaultdict(list)
    for row in con.execute(
        "select lease_id, ref_exploit, client, lessee, reference, comments from ref_fiber_lease where trim(ref_exploit) <> ''"
    ):
        fiber_by_ref[(row[1] or "").strip().upper()].append(row)

    isp_by_ref = defaultdict(list)
    for row in con.execute(
        "select lease_id, ref_exploit, client, lessee, reference, comments from ref_isp_lease where trim(ref_exploit) <> ''"
    ):
        isp_by_ref[(row[1] or "").strip().upper()].append(row)

    optical_lease_rows = list(
        con.execute(
            """
            select optical_lease_id, lease_kind, ref_exploit, reference, path_id, pair_oid,
                   feature, lessee, client, network, status, comments
            from ref_optical_lease
            """
        )
    )
    optical_lease_by_ref = defaultdict(list)
    for row in optical_lease_rows:
        for key in {(row[2] or "").strip().upper(), infer_route_ref(row[3], row[11]) or ""} - {""}:
            optical_lease_by_ref[key].append(row)

    optical_cable_rows = list(
        con.execute(
            """
            select cable_id, migration_oid, reference, userreference, comments,
                   number_of_fibers, normalized_reference, site_tokens_json
            from ref_optical_cable
            """
        )
    )
    cable_by_route = defaultdict(list)
    cable_by_site_token = defaultdict(list)
    for row in optical_cable_rows:
        for route_ref in extract_route_refs(row[2], row[3], row[4]):
            cable_by_route[route_ref].append(row)
        for token in json.loads(row[7] or "[]"):
            cable_by_site_token[token].append(row)

    optical_housing_rows = list(
        con.execute(
            """
            select housing_id, housing_type, reference, userreference, description,
                   comments, site_id, site_name
            from ref_optical_housing
            """
        )
    )
    housing_by_site = defaultdict(list)
    housing_by_route = defaultdict(list)
    for row in optical_housing_rows:
        if row[6]:
            housing_by_site[row[6]].append(row)
        for route_ref in extract_route_refs(row[2], row[3], row[4], row[5]):
            housing_by_route[route_ref].append(row)

    swag_by_service = defaultdict(list)
    for row in con.execute(
        "select interface_id, hostname, interface_name, description, route_refs_json, service_refs_json from ref_swag_interfaces"
    ):
        for service_ref in json.loads(row[5] or "[]"):
            swag_by_service[service_ref].append(row)

    network_interface_rows = list(
        con.execute(
            "select network_interface_id, device_id, device_name, source_file, source_family, interface_name, description, service_refs_json, route_refs_json, vlan_ids_json, normalized_label from ref_network_interfaces"
        )
    )
    network_interface_by_service = defaultdict(list)
    network_interface_by_route = defaultdict(list)
    for row in network_interface_rows:
        for service_ref in json.loads(row[7] or "[]"):
            network_interface_by_service[service_ref].append(row)
        for route_ref in json.loads(row[8] or "[]"):
            network_interface_by_route[route_ref].append(row)

    network_device_rows = list(
        con.execute(
            "select device_id, device_name, source_file, source_family, location_text, announcement_text, header_info, client_hint, site_hint, route_refs_json, service_refs_json, vlans_json from ref_network_devices"
        )
    )
    network_device_by_service = defaultdict(list)
    network_device_by_route = defaultdict(list)
    for row in network_device_rows:
        for service_ref in json.loads(row[10] or "[]"):
            network_device_by_service[service_ref].append(row)
        for route_ref in json.loads(row[9] or "[]"):
            network_device_by_route[route_ref].append(row)

    network_vlan_rows = list(
        con.execute(
            "select network_vlan_id, device_id, device_name, source_file, source_family, vlan_id, label, service_refs_json, route_refs_json, normalized_label from ref_network_vlans"
        )
    )

    config_rows = list(
        con.execute(
            "select config_id, file_name, hostname, header_info, service_refs_json, route_refs_json, vlans_json, normalized_payload from ref_cpe_configs"
        )
    )
    config_by_service = defaultdict(list)
    config_by_route = defaultdict(list)
    for row in config_rows:
        for service_ref in json.loads(row[4] or "[]"):
            config_by_service[service_ref].append(row)
        for route_ref in json.loads(row[5] or "[]"):
            config_by_route[route_ref].append(row)

    cpe_rows = list(con.execute("select cpe_id, hostname, normalized_hostname from ref_cpe_inventory"))

    endpoint_records = []
    optical_records = []
    network_records = []
    evidence_records = []
    service_party_records = []

    def append_optical_record(
        service_id: str,
        *,
        route_ref: str | None = None,
        route_id: str | None = None,
        route_match_rule: str | None = None,
        route_score: int | None = None,
        lease_ref: str | None = None,
        lease_id: str | None = None,
        lease_match_rule: str | None = None,
        lease_score: int | None = None,
        fiber_lease_id: str | None = None,
        fiber_lease_match_rule: str | None = None,
        fiber_lease_score: int | None = None,
        isp_lease_id: str | None = None,
        isp_lease_match_rule: str | None = None,
        isp_lease_score: int | None = None,
        support_type: str | None = None,
        support_ref: str | None = None,
        logical_route_id: str | None = None,
        cable_id: str | None = None,
        cable_match_rule: str | None = None,
        cable_score: int | None = None,
        housing_id: str | None = None,
        housing_match_rule: str | None = None,
        housing_score: int | None = None,
        site_a_optical_id: str | None = None,
        site_z_optical_id: str | None = None,
        optical_context: dict[str, object] | None = None,
    ) -> None:
        optical_records.append(
            (
                service_id,
                route_ref,
                route_id,
                route_match_rule,
                route_score,
                lease_ref,
                lease_id,
                lease_match_rule,
                lease_score,
                fiber_lease_id,
                fiber_lease_match_rule,
                fiber_lease_score,
                isp_lease_id,
                isp_lease_match_rule,
                isp_lease_score,
                support_type,
                support_ref,
                logical_route_id,
                cable_id,
                cable_match_rule,
                cable_score,
                housing_id,
                housing_match_rule,
                housing_score,
                site_a_optical_id,
                site_z_optical_id,
                json.dumps(optical_context or {}, ensure_ascii=True, sort_keys=True),
            )
        )

    services = list(
        con.execute(
            "select service_id, nature_service, principal_client, principal_external_ref, route_refs_json, service_refs_json, endpoint_a_raw, endpoint_z_raw, client_final from service_master_active"
        )
    )
    for service in services:
        service_id, nature_service, principal_client, principal_external_ref, route_refs_json, service_refs_json, endpoint_a_raw, endpoint_z_raw, client_final = service
        route_refs = json.loads(route_refs_json or "[]")
        service_refs = json.loads(service_refs_json or "[]")
        service_seeds = [client_final, endpoint_z_raw, endpoint_a_raw, principal_external_ref]

        endpoint_matches = {}
        for label, raw_value in [("A", endpoint_a_raw), ("Z", endpoint_z_raw)]:
            match = match_site(raw_value, site_index, all_sites)
            endpoint_matches[label] = match
            endpoint_records.append((service_id, label, raw_value, match.site_id, match.site_name, match.score, match.rule))
            if match.site_id:
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "site",
                        match.rule or "site_unknown",
                        match.score,
                        "ref_sites",
                        match.site_id,
                        {"endpoint_label": label, "raw_value": raw_value, "site_name": match.site_name},
                    )
                )

        contract_party_id = party_id_for_alias(party_alias_index, principal_client)
        if contract_party_id:
            service_party_records.append((service_id, "contract_party", contract_party_id, "party_exact", 100, "lea_active_lines", principal_client))
            evidence_records.append(
                build_evidence(
                    service_id,
                    "party",
                    "contract_party_exact",
                    100,
                    "party_master",
                    contract_party_id,
                    {"role": "contract_party", "label": principal_client},
                )
            )
        final_party_label = client_final or endpoint_z_raw
        final_party_id = party_id_for_alias(party_alias_index, final_party_label)
        if final_party_id:
            service_party_records.append((service_id, "final_party", final_party_id, "party_exact", 95, "lea_active_lines", final_party_label))
            evidence_records.append(
                build_evidence(
                    service_id,
                    "party",
                    "final_party_exact",
                    95,
                    "party_master",
                    final_party_id,
                    {"role": "final_party", "label": final_party_label},
                )
            )

        route_seen = set()
        site_a_id = endpoint_matches["A"].site_id
        site_z_id = endpoint_matches["Z"].site_id
        for route_ref in route_refs:
            if route_ref in route_seen:
                continue
            route_seen.add(route_ref)
            route = route_index.get(route_ref)
            logical_route = logical_route_by_ref.get(route_ref, [None])[0]
            logical_route_id = logical_route[0] if logical_route else (route[0] if route else None)
            route_id = route[0] if route else logical_route_id
            route_score = 100 if logical_route or route else 0
            append_optical_record(
                service_id,
                route_ref=route_ref,
                route_id=route_id,
                route_match_rule="route_ref_exact",
                route_score=route_score,
                support_type="route",
                support_ref=route_ref,
                logical_route_id=logical_route_id,
                site_a_optical_id=site_a_id,
                site_z_optical_id=site_z_id,
                optical_context={
                    "source_layer": logical_route[2] if logical_route else "legacy_route_ref",
                    "reference": logical_route[4] if logical_route else None,
                    "path_id": logical_route[5] if logical_route else None,
                    "pair_oid": logical_route[6] if logical_route else None,
                },
            )
            if route or logical_route:
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "optical_route",
                        "route_ref_exact",
                        100,
                        "ref_optical_logical_route" if logical_route else "ref_routes",
                        logical_route_id or route_id or route_ref,
                        {
                            "route_ref": route_ref,
                            "network": (logical_route[10] if logical_route else route[2] if route else None),
                            "client": (logical_route[9] if logical_route else route[3] if route else None),
                            "lessee": (logical_route[8] if logical_route else route[4] if route else None),
                        },
                    )
                )
            for lease in lease_by_ref.get(route_ref, []):
                append_optical_record(
                    service_id,
                    route_ref=route_ref,
                    route_id=route_id,
                    route_match_rule="route_ref_exact",
                    route_score=route_score,
                    lease_ref=route_ref,
                    lease_id=lease[0],
                    lease_match_rule="lease_ref_exact",
                    lease_score=95,
                    support_type="lease",
                    support_ref=lease[0],
                    logical_route_id=logical_route_id,
                    site_a_optical_id=site_a_id,
                    site_z_optical_id=site_z_id,
                    optical_context={"reference_l1": lease[5], "reference_l2": lease[6]},
                )
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "optical_lease",
                        "lease_ref_exact",
                        95,
                        "ref_lease_template",
                        lease[0],
                        {"route_ref": route_ref, "clients": lease[4], "lessee": lease[3], "reference_l1": lease[5], "reference_l2": lease[6]},
                    )
                )
            for fiber in fiber_by_ref.get(route_ref, []):
                append_optical_record(
                    service_id,
                    route_ref=route_ref,
                    route_id=route_id,
                    route_match_rule="route_ref_exact",
                    route_score=route_score,
                    lease_ref=route_ref,
                    fiber_lease_id=fiber[0],
                    fiber_lease_match_rule="fiber_lease_ref_exact",
                    fiber_lease_score=95,
                    support_type="fiber_lease",
                    support_ref=fiber[0],
                    logical_route_id=logical_route_id,
                    site_a_optical_id=site_a_id,
                    site_z_optical_id=site_z_id,
                    optical_context={"reference": fiber[4], "comments": fiber[5]},
                )
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "fiber_lease",
                        "fiber_lease_ref_exact",
                        95,
                        "ref_fiber_lease",
                        fiber[0],
                        {"route_ref": route_ref, "client": fiber[2], "lessee": fiber[3], "reference": fiber[4]},
                    )
                )
            for isp in isp_by_ref.get(route_ref, []):
                append_optical_record(
                    service_id,
                    route_ref=route_ref,
                    route_id=route_id,
                    route_match_rule="route_ref_exact",
                    route_score=route_score,
                    lease_ref=route_ref,
                    isp_lease_id=isp[0],
                    isp_lease_match_rule="isp_lease_ref_exact",
                    isp_lease_score=95,
                    support_type="isp_lease",
                    support_ref=isp[0],
                    logical_route_id=logical_route_id,
                    site_a_optical_id=site_a_id,
                    site_z_optical_id=site_z_id,
                    optical_context={"reference": isp[4], "comments": isp[5]},
                )
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "isp_lease",
                        "isp_lease_ref_exact",
                        95,
                        "ref_isp_lease",
                        isp[0],
                        {"route_ref": route_ref, "client": isp[2], "lessee": isp[3], "reference": isp[4]},
                    )
                )
            for interface in network_interface_by_route.get(route_ref, []):
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "route_confirmation",
                        "network_interface_route_ref_exact",
                        90,
                        "ref_network_interfaces",
                        interface[0],
                        {"route_ref": route_ref, "device_name": interface[2], "interface_name": interface[5], "description": interface[6]},
                    )
                )
            for device in network_device_by_route.get(route_ref, []):
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "route_confirmation",
                        "network_device_route_ref_exact",
                        85,
                        "ref_network_devices",
                        device[0],
                        {"route_ref": route_ref, "device_name": device[1], "announcement": device[5], "site_hint": device[8]},
                    )
                )
            for config in config_by_route.get(route_ref, []):
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "route_confirmation",
                        "config_route_ref_exact",
                        90,
                        "ref_cpe_configs",
                        config[0],
                        {"route_ref": route_ref, "hostname": config[2], "header_info": config[3], "vlans": json.loads(config[6] or "[]")},
                    )
                )
            for cable in cable_by_route.get(route_ref, []):
                cable_score = 95 if cable[5] is not None and int(cable[5]) <= 24 else 88
                append_optical_record(
                    service_id,
                    route_ref=route_ref,
                    route_id=route_id,
                    route_match_rule="route_ref_exact",
                    route_score=route_score,
                    support_type="cable",
                    support_ref=cable[0],
                    logical_route_id=logical_route_id,
                    cable_id=cable[0],
                    cable_match_rule="cable_route_ref_exact",
                    cable_score=cable_score,
                    site_a_optical_id=site_a_id,
                    site_z_optical_id=site_z_id,
                    optical_context={
                        "reference": cable[2],
                        "userreference": cable[3],
                        "number_of_fibers": cable[5],
                        "site_tokens": json.loads(cable[7] or "[]"),
                    },
                )
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "optical_cable",
                        "cable_route_ref_exact",
                        cable_score,
                        "ref_optical_cable",
                        cable[0],
                        {"route_ref": route_ref, "reference": cable[2], "number_of_fibers": cable[5]},
                    )
                )
            for housing in housing_by_route.get(route_ref, []):
                append_optical_record(
                    service_id,
                    route_ref=route_ref,
                    route_id=route_id,
                    route_match_rule="route_ref_exact",
                    route_score=route_score,
                    support_type="housing",
                    support_ref=housing[0],
                    logical_route_id=logical_route_id,
                    housing_id=housing[0],
                    housing_match_rule="housing_route_ref_exact",
                    housing_score=82,
                    site_a_optical_id=site_a_id,
                    site_z_optical_id=site_z_id,
                    optical_context={"reference": housing[2], "description": housing[4]},
                )
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "optical_housing",
                        "housing_route_ref_exact",
                        82,
                        "ref_optical_housing",
                        housing[0],
                        {"route_ref": route_ref, "reference": housing[2], "site_id": housing[6]},
                    )
                )

        if nature_service in {"IRU FON", "Location FON"} and not route_refs:
            if site_a_id and site_z_id:
                for lease, site_l1, site_l2 in lease_pair_index.get(tuple(sorted([site_a_id, site_z_id])), []):
                    route_ref = (lease[1] or "").strip().upper()
                    route = route_index.get(route_ref)
                    logical_route = logical_route_by_ref.get(route_ref, [None])[0]
                    route_id = route[0] if route else (logical_route[0] if logical_route else None)
                    append_optical_record(
                        service_id,
                        route_ref=route_ref,
                        route_id=route_id,
                        route_match_rule="lease_site_pair_match",
                        route_score=80,
                        lease_ref=route_ref,
                        lease_id=lease[0],
                        lease_match_rule="lease_site_pair_match",
                        lease_score=85,
                        support_type="lease",
                        support_ref=lease[0],
                        logical_route_id=logical_route[0] if logical_route else route_id,
                        site_a_optical_id=site_a_id,
                        site_z_optical_id=site_z_id,
                        optical_context={"reference_l1": lease[5], "reference_l2": lease[6]},
                    )
                    evidence_records.append(
                        build_evidence(
                            service_id,
                            "optical_lease",
                            "lease_site_pair_match",
                            85,
                            "ref_lease_template",
                            lease[0],
                            {"reference_l1": lease[5], "reference_l2": lease[6], "site_a": site_a_id, "site_z": site_z_id},
                        )
                    )

        if nature_service in {"IRU FON", "Location FON"}:
            site_tokens = set(extract_place_tokens(endpoint_a_raw, endpoint_z_raw, client_final, principal_external_ref))
            candidate_cables: dict[str, tuple] = {}
            for token in site_tokens:
                for cable in cable_by_site_token.get(token, []):
                    candidate_cables.setdefault(cable[0], cable)
            for cable in candidate_cables.values():
                cable_tokens = set(json.loads(cable[7] or "[]"))
                overlap = len(site_tokens & cable_tokens)
                if overlap < 1:
                    continue
                cable_score = 92 if overlap >= 2 else 80
                if cable[5] is not None and int(cable[5]) <= 24:
                    cable_score = min(96, cable_score + 4)
                append_optical_record(
                    service_id,
                    support_type="cable",
                    support_ref=cable[0],
                    cable_id=cable[0],
                    cable_match_rule="cable_site_token_overlap",
                    cable_score=cable_score,
                    site_a_optical_id=site_a_id,
                    site_z_optical_id=site_z_id,
                    optical_context={
                        "reference": cable[2],
                        "number_of_fibers": cable[5],
                        "site_tokens": sorted(cable_tokens),
                    },
                )
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "optical_cable",
                        "cable_site_token_overlap",
                        cable_score,
                        "ref_optical_cable",
                        cable[0],
                        {"reference": cable[2], "site_tokens": sorted(cable_tokens)},
                    )
                )
            if site_a_id:
                for housing in housing_by_site.get(site_a_id, []):
                    append_optical_record(
                        service_id,
                        support_type="housing",
                        support_ref=housing[0],
                        housing_id=housing[0],
                        housing_match_rule="housing_site_match",
                        housing_score=78,
                        site_a_optical_id=site_a_id,
                        site_z_optical_id=site_z_id,
                        optical_context={"reference": housing[2], "description": housing[4], "site_id": housing[6]},
                    )
            if site_z_id and site_z_id != site_a_id:
                for housing in housing_by_site.get(site_z_id, []):
                    append_optical_record(
                        service_id,
                        support_type="housing",
                        support_ref=housing[0],
                        housing_id=housing[0],
                        housing_match_rule="housing_site_match",
                        housing_score=78,
                        site_a_optical_id=site_a_id,
                        site_z_optical_id=site_z_id,
                        optical_context={"reference": housing[2], "description": housing[4], "site_id": housing[6]},
                    )

        exact_network_hit = False
        for service_ref in service_refs:
            for interface in swag_by_service.get(service_ref, []):
                exact_network_hit = True
                network_records.append((service_id, service_ref, interface[0], "service_ref_exact", 100, None, None, None, None, None, None, None, None, None, None, None, None, json.dumps([], ensure_ascii=True)))
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "network_interface",
                        "swag_service_ref_exact",
                        100,
                        "ref_swag_interfaces",
                        interface[0],
                        {"service_ref": service_ref, "hostname": interface[1], "interface_name": interface[2], "description": interface[3]},
                    )
                )
            for interface in network_interface_by_service.get(service_ref, []):
                exact_network_hit = True
                vlan_ids = json.loads(interface[9] or "[]")
                network_records.append((service_id, service_ref, None, None, None, interface[0], "network_interface_service_ref_exact", 95, None, None, None, None, None, None, None, None, None, json.dumps(vlan_ids, ensure_ascii=True)))
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "network_interface",
                        "network_interface_service_ref_exact",
                        95,
                        "ref_network_interfaces",
                        interface[0],
                        {"service_ref": service_ref, "device_name": interface[2], "interface_name": interface[5], "description": interface[6], "vlans": vlan_ids},
                    )
                )
                party_id = party_id_for_alias(party_alias_index, interface[6])
                if party_id:
                    service_party_records.append((service_id, "final_party", party_id, "network_interface_label", 85, "ref_network_interfaces", interface[0]))
            for config in config_by_service.get(service_ref, []):
                exact_network_hit = True
                inferred_vlans = json.loads(config[6] or "[]")
                cpe_match = cpe_by_hostname.get(norm_text(config[2] or ""))
                cpe_id = cpe_match[0] if cpe_match else None
                cpe_score = 90 if cpe_id else None
                network_records.append((service_id, service_ref, None, None, None, None, None, None, None, None, None, cpe_id, "config_hostname_to_cpe" if cpe_id else None, cpe_score, config[0], "config_service_ref_exact", 95, json.dumps(inferred_vlans, ensure_ascii=True)))
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "network_config",
                        "config_service_ref_exact",
                        95,
                        "ref_cpe_configs",
                        config[0],
                        {"service_ref": service_ref, "hostname": config[2], "header_info": config[3], "vlans": inferred_vlans},
                    )
                )
            for device in network_device_by_service.get(service_ref, []):
                exact_network_hit = True
                cpe_match = cpe_by_hostname.get(norm_text(device[1]))
                cpe_id = cpe_match[0] if cpe_match else None
                inferred_vlans = json.loads(device[11] or "[]")
                network_records.append((service_id, service_ref, None, None, None, None, None, None, None, None, None, cpe_id, "device_hostname_to_cpe" if cpe_id else None, 90 if cpe_id else None, None, None, None, json.dumps(inferred_vlans, ensure_ascii=True)))
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "network_device",
                        "network_device_service_ref_exact",
                        90,
                        "ref_network_devices",
                        device[0],
                        {"service_ref": service_ref, "device_name": device[1], "site_hint": device[8], "announcement": device[5]},
                    )
                )
                party_id = party_id_for_alias(party_alias_index, device[7] or device[8])
                if party_id:
                    service_party_records.append((service_id, "final_party", party_id, "network_device_hint", 80, "ref_network_devices", device[0]))

        if nature_service == "Lan To Lan":
            best_vlan = best_label_candidate(service_seeds, network_vlan_rows, 6)
            if best_vlan:
                network_vlan_id, device_id, device_name, source_file, source_family, vlan_id, label, *_rest, score = best_vlan
                network_records.append((service_id, None, None, None, None, None, None, None, network_vlan_id, "network_vlan_label_match", score, None, None, None, None, None, None, json.dumps([vlan_id], ensure_ascii=True)))
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "network_vlan",
                        "network_vlan_label_match",
                        score,
                        "ref_network_vlans",
                        network_vlan_id,
                        {"device_name": device_name, "vlan_id": vlan_id, "label": label},
                    )
                )
                party_id = party_id_for_alias(party_alias_index, label)
                if party_id:
                    service_party_records.append((service_id, "final_party", party_id, "network_vlan_label", score, "ref_network_vlans", network_vlan_id))

            best_interface = best_label_candidate(service_seeds, network_interface_rows, 6)
            if best_interface:
                (network_interface_id, device_id, device_name, source_file, source_family, interface_name, description, _service_refs_json, _route_refs_json, vlan_ids_json, _normalized_label, score) = best_interface
                network_records.append((service_id, None, None, None, None, network_interface_id, "network_interface_label_match", score, None, None, None, None, None, None, None, None, None, vlan_ids_json or json.dumps([], ensure_ascii=True)))
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "network_interface",
                        "network_interface_label_match",
                        score,
                        "ref_network_interfaces",
                        network_interface_id,
                        {"device_name": device_name, "interface_name": interface_name, "description": description, "vlans": json.loads(vlan_ids_json or "[]")},
                    )
                )
                party_id = party_id_for_alias(party_alias_index, description)
                if party_id:
                    service_party_records.append((service_id, "final_party", party_id, "network_interface_label", score, "ref_network_interfaces", network_interface_id))

            best_device = best_label_candidate(
                service_seeds,
                [
                    (
                        row[0],
                        row[1],
                        " ".join(part for part in [row[7], row[8], row[5]] if part),
                        row[7],
                        row[8],
                        row[5],
                        row[11],
                    )
                    for row in network_device_rows
                ],
                2,
            )
            if best_device:
                device_id, device_name, _combined_label, client_hint, site_hint, announcement, vlans_json, score = best_device
                cpe_match = cpe_by_hostname.get(norm_text(device_name))
                cpe_id = cpe_match[0] if cpe_match else None
                network_records.append((service_id, None, None, None, None, None, None, None, None, None, None, cpe_id, "device_label_to_cpe" if cpe_id else None, score if cpe_id else None, None, None, None, vlans_json or json.dumps([], ensure_ascii=True)))
                evidence_records.append(
                    build_evidence(
                        service_id,
                        "network_device",
                        "network_device_label_match",
                        score,
                        "ref_network_devices",
                        device_id,
                        {"device_name": device_name, "client_hint": client_hint, "site_hint": site_hint, "announcement": announcement},
                    )
                )
                party_id = party_id_for_alias(party_alias_index, client_hint or site_hint)
                if party_id:
                    service_party_records.append((service_id, "final_party", party_id, "network_device_label", score, "ref_network_devices", device_id))

            if not exact_network_hit:
                cpe_match = match_cpe(client_final or endpoint_z_raw or endpoint_a_raw, cpe_rows)
                if cpe_match:
                    cpe_id, hostname, score = cpe_match
                    network_records.append((service_id, None, None, None, None, None, None, None, None, None, None, cpe_id, "cpe_token_overlap", score, None, None, None, json.dumps([], ensure_ascii=True)))
                    evidence_records.append(
                        build_evidence(
                            service_id,
                            "cpe",
                            "cpe_token_overlap",
                            score,
                            "ref_cpe_inventory",
                            cpe_id,
                            {"hostname": hostname, "seed": client_final or endpoint_z_raw or endpoint_a_raw},
                        )
                    )

    con.executemany("insert into service_endpoint values (?,?,?,?,?,?,?)", endpoint_records)
    con.executemany("insert into service_party values (?,?,?,?,?,?,?)", service_party_records)
    con.executemany(
        """
        insert into service_support_optique (
            service_id, route_ref, route_id, route_match_rule, route_score,
            lease_ref, lease_id, lease_match_rule, lease_score,
            fiber_lease_id, fiber_lease_match_rule, fiber_lease_score,
            isp_lease_id, isp_lease_match_rule, isp_lease_score,
            support_type, support_ref, logical_route_id,
            cable_id, cable_match_rule, cable_score,
            housing_id, housing_match_rule, housing_score,
            site_a_optical_id, site_z_optical_id, optical_context_json
        ) values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        optical_records,
    )
    con.executemany("insert into service_support_reseau values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", network_records)
    con.executemany("insert into service_match_evidence values (?,?,?,?,?,?,?,?)", evidence_records)
    con.commit()
    LOG.info(
        "Created %s endpoint, %s party, %s optical, %s network links",
        len(endpoint_records),
        len(service_party_records),
        len(optical_records),
        len(network_records),
    )


def match_cpe(seed: str, cpe_rows: list[tuple]) -> tuple[str, str, int] | None:
    normalized = norm_text(seed)
    if not normalized:
        return None
    seed_tokens = {token for token in normalized.split() if len(token) >= 5}
    best = None
    best_score = 0
    for cpe_id, hostname, normalized_hostname in cpe_rows:
        host_tokens = set((normalized_hostname or "").split())
        overlap = len(seed_tokens & host_tokens)
        if overlap >= 2:
            score = min(85, overlap * 20)
            if score > best_score:
                best = (cpe_id, hostname, score)
                best_score = score
    return best


def build_evidence(
    service_id: str,
    evidence_type: str,
    rule_name: str,
    score: int,
    source_table: str,
    source_key: str,
    payload: dict[str, object],
) -> tuple[str, str, str, str, int, str, str, str]:
    evidence_id = f"EVD-{safe_hash([service_id, evidence_type, rule_name, source_key, json.dumps(payload, sort_keys=True)])}"
    return (
        evidence_id,
        service_id,
        evidence_type,
        rule_name,
        score,
        source_table,
        source_key,
        json.dumps(payload, ensure_ascii=True, sort_keys=True),
    )


def best_scored_row(rows: list[tuple], score_indexes: list[int]) -> tuple | None:
    best = None
    best_score = -1
    for row in rows:
        score = max((row[index] or 0) for index in score_indexes)
        if score > best_score:
            best = row
            best_score = score
    return best


def build_publication_views(con: sqlite3.Connection) -> None:
    LOG.info("Building publication views")
    con.execute("delete from gold_service_active")
    con.execute("delete from service_review_queue")

    party_rows = defaultdict(list)
    for row in con.execute("select service_id, role_name, party_id, rule_name, score, source_table, source_key from service_party"):
        party_rows[(row[0], row[1])].append(row)

    endpoint_rows = defaultdict(list)
    for row in con.execute("select service_id, endpoint_label, raw_value, matched_site_id, matched_site_name, score, rule_name from service_endpoint"):
        endpoint_rows[(row[0], row[1])].append(row)

    optical_rows = defaultdict(list)
    for row in con.execute("select * from service_support_optique"):
        optical_rows[row[0]].append(row)

    network_rows = defaultdict(list)
    for row in con.execute("select * from service_support_reseau"):
        network_rows[row[0]].append(row)

    evidence_stats = defaultdict(lambda: {"count": 0, "strong": 0})
    for service_id, score in con.execute("select service_id, score from service_match_evidence"):
        evidence_stats[service_id]["count"] += 1
        if (score or 0) >= 90:
            evidence_stats[service_id]["strong"] += 1

    gold_rows = []
    review_rows = []

    services = list(
        con.execute(
            "select service_id, nature_service, principal_client, principal_offer, principal_external_ref, endpoint_a_raw, endpoint_z_raw, client_final from service_master_active"
        )
    )

    for service_id, nature_service, principal_client, principal_offer, principal_external_ref, endpoint_a_raw, endpoint_z_raw, client_final in services:
        contract_party_row = best_scored_row(party_rows.get((service_id, "contract_party"), []), [4])
        final_party_row = best_scored_row(party_rows.get((service_id, "final_party"), []), [4])
        endpoint_a_row = best_scored_row(endpoint_rows.get((service_id, "A"), []), [5])
        endpoint_z_row = best_scored_row(endpoint_rows.get((service_id, "Z"), []), [5])
        optical_candidates = optical_rows.get(service_id, [])
        optical_row = None
        best_optical_score = 0
        for candidate in optical_candidates:
            candidate_scores = [candidate[index] or 0 for index in (4, 8, 11, 14)]
            if len(candidate) > 20:
                candidate_scores.append(candidate[20] or 0)
            if len(candidate) > 23:
                candidate_scores.append(candidate[23] or 0)
            score = max(candidate_scores) if candidate_scores else 0
            if optical_row is None or score > best_optical_score:
                optical_row = candidate
                best_optical_score = score
        network_row = best_scored_row(network_rows.get(service_id, []), [4, 7, 10, 13, 16])

        best_network_score = max((network_row[index] or 0) for index in [4, 7, 10, 13, 16]) if network_row else 0
        endpoint_a_score = endpoint_a_row[5] if endpoint_a_row else 0
        endpoint_z_score = endpoint_z_row[5] if endpoint_z_row else 0
        strong_evidence_count = evidence_stats[service_id]["strong"]
        evidence_count = evidence_stats[service_id]["count"]

        match_state = "review_required"
        if nature_service == "Lan To Lan":
            if best_network_score >= 95 and (endpoint_z_score >= 60 or final_party_row):
                match_state = "auto_valid"
            elif best_network_score < 72:
                match_state = "review_required"
        elif nature_service in {"IRU FON", "Location FON"}:
            if best_optical_score >= 95 and (endpoint_a_score >= 60 or endpoint_z_score >= 60):
                match_state = "auto_valid"
        elif nature_service == "Hebergement":
            if max(endpoint_a_score, endpoint_z_score) >= 90:
                match_state = "auto_valid"

        confidence_score = max(best_optical_score, best_network_score, endpoint_a_score, endpoint_z_score)
        if match_state == "auto_valid" and strong_evidence_count >= 2:
            confidence_band = "high"
        elif confidence_score >= 80:
            confidence_band = "medium"
        else:
            confidence_band = "low"

        optical_context = {}
        if optical_row and len(optical_row) > 26 and optical_row[26]:
            try:
                optical_context = json.loads(optical_row[26])
            except json.JSONDecodeError:
                optical_context = {}

        gold_rows.append(
            (
                service_id,
                match_state,
                confidence_band,
                contract_party_row[2] if contract_party_row else None,
                final_party_row[2] if final_party_row else None,
                endpoint_a_row[3] if endpoint_a_row else None,
                endpoint_z_row[3] if endpoint_z_row else None,
                optical_row[1] if optical_row else None,
                optical_row[2] if optical_row else None,
                optical_row[6] if optical_row else None,
                optical_row[9] if optical_row else None,
                optical_row[12] if optical_row else None,
                network_row[2] if network_row else None,
                network_row[5] if network_row else None,
                network_row[8] if network_row else None,
                network_row[11] if network_row else None,
                network_row[14] if network_row else None,
                network_row[17] if network_row else json.dumps([], ensure_ascii=True),
                strong_evidence_count,
                evidence_count,
                json.dumps(
                    {
                        "nature_service": nature_service,
                        "principal_client": principal_client,
                        "principal_offer": principal_offer,
                        "principal_external_ref": principal_external_ref,
                        "endpoint_a_raw": endpoint_a_raw,
                        "endpoint_z_raw": endpoint_z_raw,
                        "client_final": client_final,
                        "best_optical_score": best_optical_score,
                        "best_network_score": best_network_score,
                        "endpoint_a_score": endpoint_a_score,
                        "endpoint_z_score": endpoint_z_score,
                        "logical_route_id": optical_row[17] if optical_row and len(optical_row) > 17 else None,
                        "cable_id": optical_row[18] if optical_row and len(optical_row) > 18 else None,
                        "housing_id": optical_row[21] if optical_row and len(optical_row) > 21 else None,
                        "optical_site_a": optical_row[24] if optical_row and len(optical_row) > 24 else None,
                        "optical_site_z": optical_row[25] if optical_row and len(optical_row) > 25 else None,
                        "optical_reasoning": optical_context,
                    },
                    ensure_ascii=True,
                    sort_keys=True,
                ),
            )
        )

        issues = []
        if not contract_party_row:
            issues.append(("missing_contract_party", "high", "Contract party unresolved"))
        if not final_party_row:
            issues.append(("missing_final_party", "medium", "Final party unresolved"))
        if endpoint_z_score < 60 and nature_service != "Hebergement":
            issues.append(("missing_site_z", "medium", "Endpoint Z site unresolved"))
        if nature_service == "Lan To Lan" and best_network_score < 72:
            issues.append(("missing_network_support", "high", "No reliable network support matched"))
        if nature_service in {"IRU FON", "Location FON"} and best_optical_score < 80:
            issues.append(("missing_optical_support", "high", "No reliable optical support matched"))
        if match_state != "auto_valid":
            issues.append(("manual_review", "medium", "Service requires analyst review"))

        for review_type, severity, reason in issues:
            review_id = f"RVW-{safe_hash([service_id, review_type, reason])}"
            review_rows.append(
                (
                    review_id,
                    service_id,
                    review_type,
                    severity,
                    "open",
                    reason,
                    json.dumps(
                        {
                            "nature_service": nature_service,
                            "principal_client": principal_client,
                            "principal_offer": principal_offer,
                            "best_optical_score": best_optical_score,
                            "best_network_score": best_network_score,
                            "endpoint_a_score": endpoint_a_score,
                            "endpoint_z_score": endpoint_z_score,
                            "client_final": client_final,
                        },
                        ensure_ascii=True,
                        sort_keys=True,
                    ),
                )
            )

    con.executemany("insert into gold_service_active values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", gold_rows)
    con.executemany("insert into service_review_queue values (?,?,?,?,?,?,?)", review_rows)
    con.commit()
    LOG.info("Built %s gold services and %s review items", len(gold_rows), len(review_rows))


def _table_exists(con: sqlite3.Connection, table_name: str) -> bool:
    return (
        con.execute(
            "select name from sqlite_master where type = 'table' and name = ?",
            (table_name,),
        ).fetchone()
        is not None
    )


def _table_columns(con: sqlite3.Connection, table_name: str) -> set[str]:
    return {row[1] for row in con.execute(f'pragma table_info("{table_name}")').fetchall()}


def _get_row_value(row: sqlite3.Row, key: str, default: object = None) -> object:
    return row[key] if key in row.keys() else default


def _load_ref_site_maps(con: sqlite3.Connection) -> tuple[dict[str, str], dict[str, list[tuple[str, str]]]]:
    site_cols = _table_columns(con, "ref_sites")
    name_column = "reference" if "reference" in site_cols else "site_id"
    alt_name_column = "userreference" if "userreference" in site_cols else None
    normalized_columns = [col for col in ("normalized_reference", "normalized_userreference", "normalized_address") if col in site_cols]

    site_name_by_id: dict[str, str] = {}
    site_candidates: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for row in con.execute('select * from ref_sites').fetchall():
        site_id = row["site_id"]
        site_name = row[name_column] or site_id
        site_name_by_id[site_id] = site_name

        seen_keys = {site_id}
        site_candidates[norm_text(site_id)].append((site_id, site_name))
        candidate_values = [site_name]
        if alt_name_column:
            candidate_values.append(row[alt_name_column])
        for col in normalized_columns:
            candidate_values.append(row[col])

        for value in candidate_values:
            normalized = norm_text(value)
            if not normalized or normalized in seen_keys:
                continue
            site_candidates[normalized].append((site_id, site_name))
            seen_keys.add(normalized)

    return site_name_by_id, site_candidates


def _resolve_agent_site(
    raw_value: object,
    site_candidates: dict[str, list[tuple[str, str]]],
) -> tuple[str | None, str | None, bool]:
    normalized = norm_text(raw_value)
    if not normalized:
        return None, None, False
    matches = site_candidates.get(normalized, [])
    if len(matches) != 1:
        return None, None, bool(matches)
    site_id, site_name = matches[0]
    return site_id, site_name, True


def _party_name_by_id(con: sqlite3.Connection) -> dict[str, str]:
    return {
        row["party_id"]: row["canonical_name"]
        for row in con.execute("select party_id, canonical_name from party_master").fetchall()
    }


def _latest_agent_resolution_map(con: sqlite3.Connection) -> dict[str, sqlite3.Row]:
    if not _table_exists(con, "agent_resolutions"):
        return {}
    rows = con.execute(
        """
        select *
        from agent_resolutions
        order by service_id, created_at desc, rowid desc
        """
    ).fetchall()
    latest: dict[str, sqlite3.Row] = {}
    for row in rows:
        latest.setdefault(row["service_id"], row)
    return latest


def _latest_validated_agent_resolution_map(con: sqlite3.Connection) -> dict[str, sqlite3.Row]:
    latest = _latest_agent_resolution_map(con)
    return {
        service_id: row
        for service_id, row in latest.items()
        if row["status"] == "validated"
    }


def _best_party_map(con: sqlite3.Connection, role_name: str) -> dict[str, str]:
    rows = con.execute(
        """
        select service_id, party_id, score
        from service_party
        where role_name = ?
        order by service_id, score desc, party_id
        """,
        (role_name,),
    ).fetchall()
    best: dict[str, str] = {}
    for row in rows:
        best.setdefault(row["service_id"], row["party_id"])
    return best


def _bss_link_map(con: sqlite3.Connection) -> dict[str, dict[str, object]]:
    links: dict[str, dict[str, object]] = {}
    if not _table_exists(con, "service_bss_line"):
        return links

    rows = con.execute(
        """
        select service_id, lea_line_id, role_ligne, is_principal
        from service_bss_line
        order by service_id, is_principal desc, lea_line_id
        """
    ).fetchall()
    grouped: dict[str, list[sqlite3.Row]] = defaultdict(list)
    for row in rows:
        grouped[row["service_id"]].append(row)

    for service_id, members in grouped.items():
        principal = next((row for row in members if row["is_principal"] == 1), members[0])
        links[service_id] = {
            "principal_lea_line_id": principal["lea_line_id"],
            "lea_line_ids_json": json.dumps([row["lea_line_id"] for row in members], ensure_ascii=True),
        }
    return links


def _site_name_for_id(site_id: str | None, site_name_by_id: dict[str, str]) -> str | None:
    if not site_id:
        return None
    return site_name_by_id.get(site_id) or site_id


def _resolve_agent_optical(
    service_id: str,
    agent_optical_ref: str,
    optical_rows: dict[str, list[sqlite3.Row]],
) -> tuple[dict[str, str | None] | None, bool]:
    if not agent_optical_ref:
        return None, False
    normalized = str(agent_optical_ref).strip()
    if not normalized:
        return None, False

    for row in optical_rows.get(service_id, []):
        if normalized in {
            row["route_ref"],
            row["route_id"],
            row["lease_ref"],
            row["lease_id"],
            row["fiber_lease_id"],
            row["isp_lease_id"],
            _get_row_value(row, "cable_id"),
            _get_row_value(row, "housing_id"),
            _get_row_value(row, "support_ref"),
        }:
            return {
                "route_ref": row["route_ref"],
                "route_id": row["route_id"],
                "lease_id": row["lease_id"],
                "fiber_lease_id": row["fiber_lease_id"],
                "isp_lease_id": row["isp_lease_id"],
                "logical_route_id": _get_row_value(row, "logical_route_id"),
                "cable_id": _get_row_value(row, "cable_id"),
                "housing_id": _get_row_value(row, "housing_id"),
                "optical_site_a_id": _get_row_value(row, "site_a_optical_id"),
                "optical_site_z_id": _get_row_value(row, "site_z_optical_id"),
                "optical_context_json": _get_row_value(row, "optical_context_json"),
            }, True

    return None, False


def _parse_summary_json(row: sqlite3.Row | None) -> dict[str, object]:
    if row is None or "summary_json" not in row.keys() or not row["summary_json"]:
        return {}
    try:
        return json.loads(row["summary_json"])
    except json.JSONDecodeError:
        return {}


def _critical_gap_flags(nature_service: str, gap_flags: list[str]) -> list[str]:
    critical = {
        "missing_bss_link",
        "missing_final_party",
        "missing_site_a",
    }
    if nature_service != "Hebergement":
        critical.add("missing_site_z")
    if nature_service == "Lan To Lan":
        critical.add("missing_network_support")
    if nature_service in {"IRU FON", "Location FON", "FON"}:
        critical.add("missing_optical_support")
    return [flag for flag in gap_flags if flag in critical]


def build_facturable_publication(con: sqlite3.Connection) -> None:
    LOG.info("Building facturable publication")
    original_row_factory = con.row_factory
    con.row_factory = sqlite3.Row
    try:
        con.execute("delete from service_facturable_final")

        services = con.execute("select * from service_master_active order by service_id").fetchall()
        gold_rows = {
            row["service_id"]: row
            for row in con.execute("select * from gold_service_active").fetchall()
        }
        party_names = _party_name_by_id(con)
        site_name_by_id, site_candidates = _load_ref_site_maps(con)
        bss_links = _bss_link_map(con)
        contract_party_fallback = _best_party_map(con, "contract_party")
        latest_agent = _latest_agent_resolution_map(con)
        validated_agent = {
            service_id: row for service_id, row in latest_agent.items() if row["status"] == "validated"
        }
        optical_by_service: dict[str, list[sqlite3.Row]] = defaultdict(list)
        for row in con.execute("select * from service_support_optique").fetchall():
            optical_by_service[row["service_id"]].append(row)

        rows_to_insert = []
        published_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        for service in services:
            service_id = service["service_id"]
            gold = gold_rows.get(service_id)
            gold_summary = _parse_summary_json(gold)
            latest_res = latest_agent.get(service_id)
            validated_res = validated_agent.get(service_id)
            bss_link = bss_links.get(service_id)
            gap_flags: list[str] = []
            note_parts: list[str] = []

            contract_party_id = None
            if gold is not None:
                contract_party_id = gold["contract_party_id"]
            if not contract_party_id:
                contract_party_id = contract_party_fallback.get(service_id)
            contract_party_name = party_names.get(contract_party_id or "")

            selected_truth_source = "fallback_bss_only"
            gold_match_state = gold["match_state"] if gold is not None else None
            if validated_res is not None:
                selected_truth_source = "agent_validated"
            elif gold is not None and gold_match_state == "auto_valid":
                selected_truth_source = "gold_auto_valid"
            elif gold is not None:
                selected_truth_source = "gold_review_required"

            final_party_id = None
            final_party_source = "missing"
            if validated_res is not None and (validated_res["party_final_id"] or "").strip():
                final_party_id = validated_res["party_final_id"]
                final_party_source = "agent_validated"
                note_parts.append("party from agent")
            elif gold is not None and (gold["final_party_id"] or "").strip():
                final_party_id = gold["final_party_id"]
                final_party_source = "gold"
                if selected_truth_source == "agent_validated":
                    note_parts.append("party from gold")
            else:
                gap_flags.append("missing_final_party")
            final_party_name = party_names.get(final_party_id or "")

            site_a_id = gold["endpoint_a_site_id"] if gold is not None else None
            site_a_name = _site_name_for_id(site_a_id, site_name_by_id)
            site_a_source = "gold" if site_a_id else "missing"
            if validated_res is not None and "resolved_site_a_id" in validated_res.keys() and (validated_res["resolved_site_a_id"] or "").strip():
                site_a_id = validated_res["resolved_site_a_id"]
                site_a_name = _site_name_for_id(site_a_id, site_name_by_id)
                site_a_source = "agent_validated"
                note_parts.append("site_a from agent")
            elif validated_res is not None and (validated_res["site_a"] or "").strip():
                resolved_id, resolved_name, is_unique = _resolve_agent_site(validated_res["site_a"], site_candidates)
                if resolved_id and is_unique:
                    site_a_id = resolved_id
                    site_a_name = resolved_name
                    site_a_source = "agent_validated"
                    note_parts.append("site_a from agent")
                elif not site_a_id:
                    site_a_source = "agent_hint_unmapped"
                    gap_flags.append("agent_hint_unmapped")
            if not site_a_id:
                gap_flags.append("missing_site_a")

            site_z_id = gold["endpoint_z_site_id"] if gold is not None else None
            site_z_name = _site_name_for_id(site_z_id, site_name_by_id)
            site_z_source = "gold" if site_z_id else "missing"
            if validated_res is not None and "resolved_site_z_id" in validated_res.keys() and (validated_res["resolved_site_z_id"] or "").strip():
                site_z_id = validated_res["resolved_site_z_id"]
                site_z_name = _site_name_for_id(site_z_id, site_name_by_id)
                site_z_source = "agent_validated"
                note_parts.append("site_z from agent")
            elif validated_res is not None and (validated_res["site_z"] or "").strip():
                resolved_id, resolved_name, is_unique = _resolve_agent_site(validated_res["site_z"], site_candidates)
                if resolved_id and is_unique:
                    site_z_id = resolved_id
                    site_z_name = resolved_name
                    site_z_source = "agent_validated"
                    note_parts.append("site_z from agent")
                elif not site_z_id:
                    site_z_source = "agent_hint_unmapped"
                    gap_flags.append("agent_hint_unmapped")
            if not site_z_id:
                gap_flags.append("missing_site_z")

            route_ref = gold["route_ref"] if gold is not None else None
            route_id = gold["route_id"] if gold is not None else None
            lease_id = gold["lease_id"] if gold is not None else None
            fiber_lease_id = gold["fiber_lease_id"] if gold is not None else None
            isp_lease_id = gold["isp_lease_id"] if gold is not None else None
            logical_route_id = gold_summary.get("logical_route_id")
            cable_id = gold_summary.get("cable_id")
            housing_id = gold_summary.get("housing_id")
            optical_site_a_id = gold_summary.get("optical_site_a")
            optical_site_z_id = gold_summary.get("optical_site_z")
            optical_context = gold_summary.get("optical_reasoning") or {}
            optical_source = "gold" if any((route_ref, route_id, lease_id, fiber_lease_id, isp_lease_id)) else "missing"
            agent_optical_support_hint = None
            if validated_res is not None and any(
                (
                    ("route_ref" in validated_res.keys() and (validated_res["route_ref"] or "").strip()),
                    ("route_id" in validated_res.keys() and (validated_res["route_id"] or "").strip()),
                    ("lease_id" in validated_res.keys() and (validated_res["lease_id"] or "").strip()),
                    ("fiber_lease_id" in validated_res.keys() and (validated_res["fiber_lease_id"] or "").strip()),
                    ("isp_lease_id" in validated_res.keys() and (validated_res["isp_lease_id"] or "").strip()),
                    ("cable_id" in validated_res.keys() and (validated_res["cable_id"] or "").strip()),
                    ("housing_id" in validated_res.keys() and (validated_res["housing_id"] or "").strip()),
                )
            ):
                route_ref = validated_res["route_ref"] if "route_ref" in validated_res.keys() else route_ref
                route_id = validated_res["route_id"] if "route_id" in validated_res.keys() else route_id
                lease_id = validated_res["lease_id"] if "lease_id" in validated_res.keys() else lease_id
                fiber_lease_id = validated_res["fiber_lease_id"] if "fiber_lease_id" in validated_res.keys() else fiber_lease_id
                isp_lease_id = validated_res["isp_lease_id"] if "isp_lease_id" in validated_res.keys() else isp_lease_id
                logical_route_id = validated_res["route_id"] if "route_id" in validated_res.keys() and validated_res["route_id"] else logical_route_id
                cable_id = validated_res["cable_id"] if "cable_id" in validated_res.keys() else cable_id
                housing_id = validated_res["housing_id"] if "housing_id" in validated_res.keys() else housing_id
                optical_source = "agent_validated"
                note_parts.append("optical from agent")
            elif validated_res is not None and (validated_res["optical_support_ref"] or "").strip():
                agent_optical_support_hint = validated_res["optical_support_ref"]
                agent_optical_ref = validated_res["optical_support_ref"].strip()
                remapped = None
                for row in optical_by_service.get(service_id, []):
                    if agent_optical_ref in {
                        row["route_ref"],
                        row["route_id"],
                        row["lease_ref"],
                        row["lease_id"],
                        row["fiber_lease_id"],
                        row["isp_lease_id"],
                    }:
                        remapped = row
                        break
                if remapped is not None:
                    route_ref = remapped["route_ref"]
                    route_id = remapped["route_id"]
                    lease_id = remapped["lease_id"]
                    fiber_lease_id = remapped["fiber_lease_id"]
                    isp_lease_id = remapped["isp_lease_id"]
                    logical_route_id = _get_row_value(remapped, "logical_route_id")
                    cable_id = _get_row_value(remapped, "cable_id")
                    housing_id = _get_row_value(remapped, "housing_id")
                    optical_site_a_id = _get_row_value(remapped, "site_a_optical_id")
                    optical_site_z_id = _get_row_value(remapped, "site_z_optical_id")
                    if _get_row_value(remapped, "optical_context_json"):
                        try:
                            optical_context = json.loads(_get_row_value(remapped, "optical_context_json") or "{}")
                        except json.JSONDecodeError:
                            optical_context = {}
                    optical_source = "agent_validated"
                    note_parts.append("optical from agent")
                elif not any((route_ref, route_id, lease_id, fiber_lease_id, isp_lease_id, cable_id, housing_id)):
                    optical_source = "agent_hint_unmapped"
                    gap_flags.append("agent_hint_unmapped")
            if not any((route_ref, route_id, lease_id, fiber_lease_id, isp_lease_id, cable_id, housing_id)):
                gap_flags.append("missing_optical_support")
            if not any((route_ref, route_id, lease_id, fiber_lease_id, isp_lease_id)):
                gap_flags.append("missing_logical_route_ref")

            interface_id = gold["interface_id"] if gold is not None else None
            network_interface_id = gold["network_interface_id"] if gold is not None else None
            network_vlan_id = gold["network_vlan_id"] if gold is not None else None
            cpe_id = gold["cpe_id"] if gold is not None else None
            config_id = gold["config_id"] if gold is not None else None
            inferred_vlans_json = gold["inferred_vlans_json"] if gold is not None else json.dumps([], ensure_ascii=True)
            network_source = (
                "gold"
                if any((interface_id, network_interface_id, network_vlan_id, cpe_id, config_id))
                else "missing"
            )
            agent_network_support_hint = None
            if validated_res is not None and any(
                (
                    ("network_interface_id" in validated_res.keys() and (validated_res["network_interface_id"] or "").strip()),
                    ("network_vlan_id" in validated_res.keys() and (validated_res["network_vlan_id"] or "").strip()),
                    ("cpe_id" in validated_res.keys() and (validated_res["cpe_id"] or "").strip()),
                    ("config_id" in validated_res.keys() and (validated_res["config_id"] or "").strip()),
                    ("inferred_vlans_json" in validated_res.keys() and (validated_res["inferred_vlans_json"] or "").strip()),
                )
            ):
                network_interface_id = validated_res["network_interface_id"] if "network_interface_id" in validated_res.keys() else network_interface_id
                network_vlan_id = validated_res["network_vlan_id"] if "network_vlan_id" in validated_res.keys() else network_vlan_id
                cpe_id = validated_res["cpe_id"] if "cpe_id" in validated_res.keys() else cpe_id
                config_id = validated_res["config_id"] if "config_id" in validated_res.keys() else config_id
                inferred_vlans_json = validated_res["inferred_vlans_json"] if "inferred_vlans_json" in validated_res.keys() and validated_res["inferred_vlans_json"] else inferred_vlans_json
                network_source = "agent_validated"
                note_parts.append("network from agent")
            elif validated_res is not None and (validated_res["network_support_id"] or "").strip():
                agent_network_support_hint = validated_res["network_support_id"]
                if not any((interface_id, network_interface_id, network_vlan_id, cpe_id, config_id)):
                    network_source = "agent_hint_unmapped"
                    gap_flags.append("agent_hint_unmapped")
            if not any((interface_id, network_interface_id, network_vlan_id, cpe_id, config_id)):
                gap_flags.append("missing_network_support")

            if bss_link is None:
                gap_flags.append("missing_bss_link")
                principal_lea_line_id = None
                lea_line_ids_json = json.dumps([], ensure_ascii=True)
            else:
                principal_lea_line_id = bss_link["principal_lea_line_id"]
                lea_line_ids_json = bss_link["lea_line_ids_json"]

            unique_gap_flags = sorted(set(gap_flags))
            critical_gaps = _critical_gap_flags(service["nature_service"], unique_gap_flags)

            if selected_truth_source == "agent_validated" and not critical_gaps:
                publication_status = "published_validated"
            elif selected_truth_source in {"gold_auto_valid", "gold_review_required"} and not critical_gaps:
                publication_status = "published_from_gold"
            elif critical_gaps:
                publication_status = "needs_review"
            else:
                publication_status = "published_with_gaps"

            if selected_truth_source.startswith("gold") and publication_status == "published_with_gaps" and not note_parts:
                note_parts.append("published from gold")
            if selected_truth_source == "fallback_bss_only" and not note_parts:
                note_parts.append("bss only")
            if selected_truth_source == "agent_validated" and not note_parts:
                note_parts.append("validated agent override")

            publication_comment = (
                f"source={selected_truth_source}; "
                f"gaps={','.join(unique_gap_flags) if unique_gap_flags else 'none'}; "
                f"note={', '.join(note_parts) if note_parts else 'aligned'}"
            )

            rows_to_insert.append(
                (
                    service_id,
                    service["service_key"],
                    principal_lea_line_id,
                    lea_line_ids_json,
                    service["line_count"],
                    service["active_line_count"],
                    service["nature_service"],
                    service["principal_offer"],
                    service["principal_external_ref"],
                    service["principal_internal_ref"],
                    service["principal_client"],
                    service["client_final"],
                    contract_party_id,
                    contract_party_name,
                    final_party_id,
                    final_party_name,
                    final_party_source,
                    service["endpoint_a_raw"],
                    service["endpoint_z_raw"],
                    site_a_id,
                    site_a_name,
                    site_a_source,
                    site_z_id,
                    site_z_name,
                    site_z_source,
                    route_ref,
                    route_id,
                    lease_id,
                    fiber_lease_id,
                    isp_lease_id,
                    logical_route_id,
                    cable_id,
                    housing_id,
                    optical_site_a_id,
                    optical_site_z_id,
                    json.dumps(optical_context, ensure_ascii=True, sort_keys=True),
                    optical_source,
                    agent_optical_support_hint,
                    interface_id,
                    network_interface_id,
                    network_vlan_id,
                    cpe_id,
                    config_id,
                    inferred_vlans_json or json.dumps([], ensure_ascii=True),
                    network_source,
                    agent_network_support_hint,
                    gold_match_state,
                    gold["confidence_band"] if gold is not None else None,
                    latest_res["status"] if latest_res is not None else None,
                    latest_res["confidence"] if latest_res is not None else None,
                    publication_status,
                    selected_truth_source,
                    json.dumps(unique_gap_flags, ensure_ascii=True),
                    publication_comment,
                    published_at,
                )
            )

        placeholders = ", ".join("?" for _ in range(55))
        con.executemany(
            f"insert into service_facturable_final values ({placeholders})",
            rows_to_insert,
        )
        con.commit()
        LOG.info("Built %s final facturable rows", len(rows_to_insert))
    finally:
        con.row_factory = original_row_factory


def export_csv(con: sqlite3.Connection, query: str, path: Path) -> None:
    rows = con.execute(query)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow([column[0] for column in rows.description])
        writer.writerows(rows.fetchall())


def build_report(con: sqlite3.Connection) -> None:
    LOG.info("Building report")
    total_active = con.execute("select count(*) from lea_active_lines").fetchone()[0]
    total_services = con.execute("select count(*) from service_master_active").fetchone()[0]
    by_nature = con.execute(
        "select nature_service, count(*) from service_master_active group by nature_service order by count(*) desc"
    ).fetchall()
    route_matched = con.execute(
        "select count(distinct service_id) from service_support_optique where route_id is not null"
    ).fetchone()[0]
    lease_matched = con.execute(
        "select count(distinct service_id) from service_support_optique where lease_id is not null or fiber_lease_id is not null or isp_lease_id is not null"
    ).fetchone()[0]
    cable_matched = con.execute(
        "select count(distinct service_id) from service_support_optique where cable_id is not null"
    ).fetchone()[0]
    housing_matched = con.execute(
        "select count(distinct service_id) from service_support_optique where housing_id is not null"
    ).fetchone()[0]
    network_matched = con.execute(
        "select count(distinct service_id) from service_support_reseau where interface_id is not null or network_interface_id is not null or network_vlan_id is not null or cpe_id is not null or config_id is not null"
    ).fetchone()[0]
    site_matched = con.execute(
        "select count(distinct service_id) from service_endpoint where matched_site_id is not null"
    ).fetchone()[0]
    auto_candidates = con.execute(
        "select count(distinct service_id) from service_match_evidence where score >= 95 and evidence_type <> 'party'"
    ).fetchone()[0]
    auto_valid = con.execute(
        "select count(*) from gold_service_active where match_state = 'auto_valid'"
    ).fetchone()[0]
    review_required = con.execute(
        "select count(*) from gold_service_active where match_state = 'review_required'"
    ).fetchone()[0]
    open_reviews = con.execute(
        "select count(*) from service_review_queue where current_state = 'open'"
    ).fetchone()[0]
    party_count = con.execute("select count(*) from party_master").fetchone()[0]
    network_devices = con.execute("select count(*) from ref_network_devices").fetchone()[0]
    network_vlans = con.execute("select count(*) from ref_network_vlans").fetchone()[0]
    final_rows = con.execute("select count(*) from service_facturable_final").fetchone()[0] if _table_exists(con, "service_facturable_final") else 0
    final_status_rows = (
        con.execute(
            "select publication_status, count(*) from service_facturable_final group by publication_status order by publication_status"
        ).fetchall()
        if final_rows
        else []
    )
    final_with_party = (
        con.execute("select count(*) from service_facturable_final where final_party_id is not null and trim(final_party_id) <> ''").fetchone()[0]
        if final_rows
        else 0
    )
    final_with_site_a = (
        con.execute("select count(*) from service_facturable_final where site_a_id is not null and trim(site_a_id) <> ''").fetchone()[0]
        if final_rows
        else 0
    )
    final_with_site_z = (
        con.execute("select count(*) from service_facturable_final where site_z_id is not null and trim(site_z_id) <> ''").fetchone()[0]
        if final_rows
        else 0
    )
    final_with_network = (
        con.execute(
            """
            select count(*) from service_facturable_final
            where coalesce(interface_id, network_interface_id, network_vlan_id, cpe_id, config_id) is not null
            """
        ).fetchone()[0]
        if final_rows
        else 0
    )
    final_with_optical = (
        con.execute(
            """
            select count(*) from service_facturable_final
            where coalesce(route_ref, route_id, lease_id, fiber_lease_id, isp_lease_id, cable_id, housing_id) is not null
            """
        ).fetchone()[0]
        if final_rows
        else 0
    )
    final_truth_rows = (
        con.execute(
            "select selected_truth_source, count(*) from service_facturable_final group by selected_truth_source order by selected_truth_source"
        ).fetchall()
        if final_rows
        else []
    )

    report = [
        "# Active service referential build",
        "",
        "## Scope",
        f"- Active LEA lines loaded: {total_active}",
        f"- Active service masters built: {total_services}",
        "",
        "## Coverage",
        f"- Services with matched site: {site_matched}",
        f"- Services with matched route: {route_matched}",
        f"- Services with matched optical lease: {lease_matched}",
        f"- Services with matched optical cable: {cable_matched}",
        f"- Services with matched optical housing: {housing_matched}",
        f"- Services with network support evidence: {network_matched}",
        f"- Services with at least one strong evidence (score >= 95): {auto_candidates}",
        f"- Services auto-validated in Gold: {auto_valid}",
        f"- Services requiring review in Gold: {review_required}",
        f"- Open review queue items: {open_reviews}",
        "",
        "## Service mix",
    ]
    for nature, count in by_nature:
        report.append(f"- {nature}: {count}")
    report.extend(
        [
            "",
            "## Referential assets",
            f"- Parties in party master: {party_count}",
            f"- Parsed network devices: {network_devices}",
            f"- Parsed network vlan labels: {network_vlans}",
            "",
            "## Final facturable publication",
            f"- Published rows: {final_rows}",
            f"- Rows with final party: {final_with_party}",
            f"- Rows with site A: {final_with_site_a}",
            f"- Rows with site Z: {final_with_site_z}",
            f"- Rows with network support: {final_with_network}",
            f"- Rows with optical support: {final_with_optical}",
            "",
            "## Notes",
            "- Optical matching is built directly from the GDB: logical refs (`TOIP`, `00FT`, `FREE`) plus physical cables and housings.",
            "- Site matching uses exact aliases, addresses and token overlap on Hubsite names.",
            "- Network support uses exact SWAG/config refs first, then parsed RANCID VLAN/interface labels and CPE hints.",
            "- Gold and review queue are materialized in SQLite for immediate exploitation.",
            "- `service_facturable_final` is the published billing referential, built with priority `agent_validated > gold > review`.",
        ]
    )
    if final_status_rows:
        report.append("")
        report.append("## Final publication statuses")
        for status, count in final_status_rows:
            report.append(f"- {status}: {count}")
    if final_truth_rows:
        report.append("")
        report.append("## Final truth sources")
        for source, count in final_truth_rows:
            report.append(f"- {source}: {count}")
    (OUT_DIR / "service_referential_report.md").write_text("\n".join(report), encoding="utf-8")


def export_outputs(con: sqlite3.Connection) -> None:
    export_csv(
        con,
        """
        select g.service_id, s.nature_service, g.match_state, g.confidence_band,
               s.principal_client, s.principal_offer, s.principal_external_ref,
               s.endpoint_a_raw, s.endpoint_z_raw, s.client_final,
               g.contract_party_id, g.final_party_id,
               g.endpoint_a_site_id, g.endpoint_z_site_id,
               g.route_ref, g.route_id, g.lease_id, g.fiber_lease_id, g.isp_lease_id,
               g.interface_id, g.network_interface_id, g.network_vlan_id,
               g.cpe_id, g.config_id, g.inferred_vlans_json,
               g.strong_evidence_count, g.evidence_count
        from gold_service_active g
        join service_master_active s on s.service_id = g.service_id
        order by s.nature_service, s.principal_client, g.service_id
        """,
        OUT_DIR / "service_master_active.csv",
    )
    export_csv(
        con,
        """
        select *
        from service_facturable_final
        order by nature_service, principal_client_raw, service_id
        """,
        OUT_DIR / "service_facturable_final.csv",
    )
    export_csv(
        con,
        "select * from service_match_evidence order by service_id, evidence_type, score desc",
        OUT_DIR / "service_match_evidence.csv",
    )
    export_csv(
        con,
        "select * from service_review_queue order by severity desc, service_id, review_type",
        OUT_DIR / "service_review_queue.csv",
    )
    export_csv(
        con,
        "select * from party_master order by canonical_name",
        OUT_DIR / "party_master.csv",
    )
    export_csv(
        con,
        "select * from ref_network_vlans order by device_name, vlan_id",
        OUT_DIR / "network_vlan_catalog.csv",
    )


def main() -> None:
    setup_logging()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists():
        DB_PATH.unlink()

    con = sqlite3.connect(DB_PATH)
    create_schema(con)
    load_lea_active(con)
    load_sites(con)
    load_routes(con)
    load_lease_tables(con)
    load_swag_interfaces(con)
    load_cpe_inventory(con)
    load_cpe_configs(con)
    load_network_text_artifacts(con)
    build_party_master(con)
    build_service_master(con)
    reconcile_services(con)
    build_publication_views(con)
    build_facturable_publication(con)
    export_outputs(con)
    build_report(con)
    con.close()
    LOG.info("Build completed. Outputs available in %s", OUT_DIR)


if __name__ == "__main__":
    main()
